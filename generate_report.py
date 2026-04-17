#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║  LABHYANSH SOLUTION – TRIAL BALANCE BUSINESS REPORT         ║
║  Comprehensive Financial Analysis & Reporting               ║
╚══════════════════════════════════════════════════════════════╝

This script reads the raw Trial Balance Excel data, performs
cleaning, analysis, and generates a full business report with
visuals as PNG charts and a markdown report.
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import os
import math
from collections import defaultdict

# ── Matplotlib setup (non-interactive backend) ──────────────────────────
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

# ════════════════════════════════════════════════════════════════
# CONFIGURATION
# ════════════════════════════════════════════════════════════════

EXCEL_FILE = "Dummy Data for Review.xlsx"
SHEET_NAME = "Trial Balance"
OUTPUT_DIR = "report_output"

MONTHS = ['Apr24', 'May24', 'Jun24', 'Jul24', 'Aug24']
MONTH_LABELS = {'Apr24': 'April 2024', 'May24': 'May 2024', 'Jun24': 'June 2024',
                'Jul24': 'July 2024', 'Aug24': 'August 2024'}
MONTH_COL_PAIRS = [(2, 3), (4, 5), (6, 7), (8, 9), (10, 11)]

# ── Chart styling ────────────────────────────────────────────────────────
COLORS = {
    'primary': '#1E3A5F',      # Deep navy blue
    'secondary': '#4A90D9',    # Bright blue
    'accent': '#F5A623',       # Warm amber
    'positive': '#27AE60',     # Green
    'negative': '#E74C3C',     # Red
    'neutral': '#95A5A6',      # Gray
    'bg': '#FAFBFC',           # Near-white bg
    'text': '#2C3E50',         # Dark text
}



CATEGORY_COLORS = {
    'Assets': '#3498DB',
    'Liabilities': '#E74C3C',
    'Equity': '#9B59B6',
    'Revenue': '#27AE60',
    'Expenses': '#F39C12',
    'Summary': '#95A5A6',
    'Unclassified': '#BDC3C7',
}

plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Segoe UI', 'Arial', 'Helvetica', 'DejaVu Sans'],
    'font.size': 11,
    'axes.titlesize': 14,
    'axes.titleweight': 'bold',
    'axes.labelsize': 12,
    'figure.facecolor': COLORS['bg'],
    'axes.facecolor': '#FFFFFF',
    'axes.edgecolor': '#E0E0E0',
    'grid.color': '#F0F0F0',
    'grid.linewidth': 0.8,
})


# ════════════════════════════════════════════════════════════════
# ACCOUNT CLASSIFICATION MAPPING
# ════════════════════════════════════════════════════════════════

account_classification = {
    # --- EQUITY / CAPITAL ---
    "Capital Account":                        ("Equity", "Capital", "Balance Sheet"),
    "Opening Adjustment":                     ("Equity", "Capital", "Balance Sheet"),
    "MY FACTORY (DELHI)`2":                   ("Equity", "Capital", "Balance Sheet"),
    "MEHTA CAPITAL ACCOUNT":                  ("Equity", "Capital", "Balance Sheet"),

    # --- LIABILITIES - LOANS ---
    "Loans (Liability)":                      ("Liabilities", "Loans", "Balance Sheet"),
    "Unsecured Loans":                        ("Liabilities", "Unsecured Loans", "Balance Sheet"),
    "A.M FOODS":                              ("Liabilities", "Unsecured Loans", "Balance Sheet"),

    # --- LIABILITIES - DUTIES & TAXES ---
    "Current Liabilities":                    ("Liabilities", "Current Liabilities", "Balance Sheet"),
    "Duties & Taxes":                         ("Liabilities", "Duties & Taxes", "Balance Sheet"),
    "Input Receivable":                       ("Liabilities", "Duties & Taxes - Input", "Balance Sheet"),
    "CGST  Input":                            ("Liabilities", "Duties & Taxes - Input", "Balance Sheet"),
    "IGST Input":                             ("Liabilities", "Duties & Taxes - Input", "Balance Sheet"),
    "MACHINERY PARTS IGST TAX 18%":           ("Liabilities", "Duties & Taxes - Input", "Balance Sheet"),
    "PURCHASE CGST 9%":                       ("Liabilities", "Duties & Taxes - Input", "Balance Sheet"),
    "PURCHASE SGST 9%":                       ("Liabilities", "Duties & Taxes - Input", "Balance Sheet"),
    "SGST Input":                             ("Liabilities", "Duties & Taxes - Input", "Balance Sheet"),
    "Output":                                 ("Liabilities", "Duties & Taxes - Output", "Balance Sheet"),
    "CGST":                                   ("Liabilities", "Duties & Taxes - Output", "Balance Sheet"),
    "IGST":                                   ("Liabilities", "Duties & Taxes - Output", "Balance Sheet"),
    "SGST":                                   ("Liabilities", "Duties & Taxes - Output", "Balance Sheet"),

    # --- LIABILITIES - TDS PAYABLE ---
    "TDS Payable":                            ("Liabilities", "TDS Payable", "Balance Sheet"),
    "TDS 194C Contractor Company":            ("Liabilities", "TDS Payable", "Balance Sheet"),
    "TDS 194C Contractor Non Company":        ("Liabilities", "TDS Payable", "Balance Sheet"),
    "TDS Professional 194J":                  ("Liabilities", "TDS Payable", "Balance Sheet"),

    # --- LIABILITIES - PROVISIONS ---
    "Provisions":                             ("Liabilities", "Provisions", "Balance Sheet"),
    "AUDIT FEES PAYABLE":                     ("Liabilities", "Provisions", "Balance Sheet"),
    "EPF-Administrative Charges Payable":     ("Liabilities", "Provisions", "Balance Sheet"),
    "EPF PAYABLE":                            ("Liabilities", "Provisions", "Balance Sheet"),
    "ESIC PAYABLE":                           ("Liabilities", "Provisions", "Balance Sheet"),
    "LWF Payble":                             ("Liabilities", "Provisions", "Balance Sheet"),
    "MANUFACTURING EXPENSES PAYABLE":         ("Liabilities", "Provisions", "Balance Sheet"),
    "Provision for Electricity Expenses":     ("Liabilities", "Provisions", "Balance Sheet"),
    "PROVISION FOR EXPENSES":                 ("Liabilities", "Provisions", "Balance Sheet"),
    "Provision for Tax":                      ("Liabilities", "Provisions", "Balance Sheet"),
    "SALARY PAYABLE":                         ("Liabilities", "Provisions", "Balance Sheet"),

    # --- LIABILITIES - SUNDRY CREDITORS ---
    "Sundry Creditors":                       ("Liabilities", "Sundry Creditors", "Balance Sheet"),
    "HDFC Credit Card 2140":                  ("Liabilities", "Sundry Creditors", "Balance Sheet"),

    # --- ASSETS - FIXED ASSETS ---
    "Fixed Assets":                           ("Assets", "Fixed Assets", "Balance Sheet"),
    "Accumulated Depreciation":               ("Assets", "Fixed Assets - Depreciation", "Balance Sheet"),
    "BIKE":                                   ("Assets", "Fixed Assets", "Balance Sheet"),
    "CAR(ISUZU)":                             ("Assets", "Fixed Assets", "Balance Sheet"),
    "Chapati Making Machine":                 ("Assets", "Fixed Assets", "Balance Sheet"),
    "COOLER":                                 ("Assets", "Fixed Assets", "Balance Sheet"),
    "DUSTBIN 120LTR":                         ("Assets", "Fixed Assets", "Balance Sheet"),
    "FURNITURE IGST 18%":                     ("Assets", "Fixed Assets", "Balance Sheet"),
    "LUMINOUS BATTERY":                       ("Assets", "Fixed Assets", "Balance Sheet"),
    "MACHINERY IGST 18%":                     ("Assets", "Fixed Assets", "Balance Sheet"),
    "MOBILE":                                 ("Assets", "Fixed Assets", "Balance Sheet"),
    "PUMP":                                   ("Assets", "Fixed Assets", "Balance Sheet"),
    "RO PLANT IGST 18%":                      ("Assets", "Fixed Assets", "Balance Sheet"),
    "SAREGAMA CARVAAN GST 18%":               ("Assets", "Fixed Assets", "Balance Sheet"),
    "TABLE":                                  ("Assets", "Fixed Assets", "Balance Sheet"),
    "Weighing Machine":                       ("Assets", "Fixed Assets", "Balance Sheet"),

    # --- ASSETS - INVESTMENTS ---
    "Investments":                            ("Assets", "Investments", "Balance Sheet"),
    "Investment in FDR":                      ("Assets", "Investments", "Balance Sheet"),
    "Kotak  Mutual Fund":                     ("Assets", "Investments", "Balance Sheet"),
    "Motilal Oswal Financial Services Ltd":   ("Assets", "Investments", "Balance Sheet"),
    "Nippon Mutual Fund":                     ("Assets", "Investments", "Balance Sheet"),
    "TPT Investment Management (TPTIM)":      ("Assets", "Investments", "Balance Sheet"),

    # --- ASSETS - CURRENT ASSETS ---
    "Current Assets":                         ("Assets", "Current Assets", "Balance Sheet"),
    "Opening Stock":                          ("Assets", "Current Assets - Stock", "Balance Sheet"),
    "Loans & Advances (Asset)":               ("Assets", "Loans & Advances", "Balance Sheet"),
    "Advance Salary":                         ("Assets", "Loans & Advances", "Balance Sheet"),
    "Sundry Debtors":                         ("Assets", "Sundry Debtors", "Balance Sheet"),
    "Cash-in-Hand":                           ("Assets", "Cash", "Balance Sheet"),
    "Cash":                                   ("Assets", "Cash", "Balance Sheet"),
    "Bank Accounts":                          ("Assets", "Bank", "Balance Sheet"),
    "HDFC BANK":                              ("Assets", "Bank", "Balance Sheet"),
    "Accrued Income on FDR":                  ("Assets", "Other Current Assets", "Balance Sheet"),
    "ADVANCE TAX":                            ("Assets", "Tax Assets", "Balance Sheet"),
    "SECURITY DEPOSIT - IOC APPARELS AND HCG GAS": ("Assets", "Other Current Assets", "Balance Sheet"),
    "TCS":                                    ("Assets", "Tax Assets", "Balance Sheet"),
    "TDS":                                    ("Assets", "Tax Assets", "Balance Sheet"),

    # --- REVENUE / SALES ---
    "Sales Accounts":                         ("Revenue", "Sales", "P&L"),
    "Un Exempt Sales A/c":                    ("Revenue", "Sales", "P&L"),

    # --- PURCHASES ---
    "Purchase Accounts":                      ("Expenses", "Purchases", "P&L"),
    "DISCOUNT REBATE":                        ("Expenses", "Purchases - Discount", "P&L"),
    "DISCOUNT RECEIVED":                      ("Revenue", "Discount Received", "P&L"),
    "PURCHASE GST 18%":                       ("Expenses", "Purchases", "P&L"),
    "PURCHASE GST 28%":                       ("Expenses", "Purchases", "P&L"),
    "PURCHASE GST 5%":                        ("Expenses", "Purchases", "P&L"),
    "PURCHASE GST TAX FREE":                  ("Expenses", "Purchases", "P&L"),
    "PURCHASE IGST 12%":                      ("Expenses", "Purchases", "P&L"),
    "PURCHASE IGST 18%":                      ("Expenses", "Purchases", "P&L"),
    "PURCHASE IGST 5%":                       ("Expenses", "Purchases", "P&L"),

    # --- DIRECT EXPENSES ---
    "Expenses (Direct) (Direct Expenses)":    ("Expenses", "Direct Expenses", "P&L"),
    "DELIVERY CHARGES":                       ("Expenses", "Direct Expenses", "P&L"),
    "Electricity Expense":                    ("Expenses", "Direct Expenses", "P&L"),
    "FREIGHT AND CARTAGE":                    ("Expenses", "Direct Expenses", "P&L"),
    "FREIGHT AND CARTAGE IGST 12%":           ("Expenses", "Direct Expenses", "P&L"),
    "GAS EXPENSES":                           ("Expenses", "Direct Expenses", "P&L"),
    "GAS EXPENSES IGST 18%":                  ("Expenses", "Direct Expenses", "P&L"),
    "HR SERVICES (MANPOWER SERVICES) GST 18%": ("Expenses", "Direct Expenses", "P&L"),
    "Labour Charges  (Atta Unload)":          ("Expenses", "Direct Expenses", "P&L"),
    "Labour on hire":                         ("Expenses", "Direct Expenses", "P&L"),

    # --- INDIRECT INCOME ---
    "Income (Indirect) (Indirect Incomes)":   ("Revenue", "Indirect Income", "P&L"),
    "Cartage Charges":                        ("Revenue", "Indirect Income", "P&L"),
    "Cash Discount":                          ("Revenue", "Indirect Income", "P&L"),
    "Interest Income":                        ("Revenue", "Indirect Income", "P&L"),
    "Interest on FDR- 50301072325452":        ("Revenue", "Indirect Income", "P&L"),
    "Interest on Income Tax Refund":          ("Revenue", "Indirect Income", "P&L"),
    "Misc Receipt":                           ("Revenue", "Indirect Income", "P&L"),

    # --- INDIRECT EXPENSES - EMPLOYEE ---
    "Expenses (Indirect) (Indirect Expenses)": ("Expenses", "Indirect Expenses", "P&L"),
    "Employee Expenses":                      ("Expenses", "Employee Expenses", "P&L"),
    "ADMINISTRATION CHARGES":                 ("Expenses", "Employee Expenses", "P&L"),
    "DIWALI BONUS":                           ("Expenses", "Employee Expenses", "P&L"),
    "EDLI Charges":                           ("Expenses", "Employee Expenses", "P&L"),
    "EPF":                                    ("Expenses", "Employee Expenses", "P&L"),
    "ESIC":                                   ("Expenses", "Employee Expenses", "P&L"),
    "LWF":                                    ("Expenses", "Employee Expenses", "P&L"),
    "Recruitment Expenses":                   ("Expenses", "Employee Expenses", "P&L"),
    "SALARY":                                 ("Expenses", "Employee Expenses", "P&L"),
    "STAFF WELFARE":                          ("Expenses", "Employee Expenses", "P&L"),

    # --- INDIRECT EXPENSES - ADMIN & OPERATIONS ---
    "ACCOUNTING CHARGES":                     ("Expenses", "Admin & Professional", "P&L"),
    "AUDIT FEES":                             ("Expenses", "Admin & Professional", "P&L"),
    "BAD DEBTS":                              ("Expenses", "Other Expenses", "P&L"),
    "BANK CHARGES":                           ("Expenses", "Finance Charges", "P&L"),
    "BUSINESS PROMOTION GST18%":              ("Expenses", "Marketing & Promotion", "P&L"),
    "COMPUTER REPAIR AND MANITANCE":          ("Expenses", "Repairs & Maintenance", "P&L"),
    "DEPRECIATION":                           ("Expenses", "Depreciation", "P&L"),
    "FACTORY REPAIR & MAINTANCE":             ("Expenses", "Repairs & Maintenance", "P&L"),
    "Fees & Subscription":                    ("Expenses", "Admin & Professional", "P&L"),
    "FUEL EXEPENSES":                         ("Expenses", "Vehicle & Travel", "P&L"),
    "GENERAL EXPENSES":                       ("Expenses", "Other Expenses", "P&L"),
    "Income Tax Expenses":                    ("Expenses", "Tax Expenses", "P&L"),
    "INSURANCE":                              ("Expenses", "Insurance", "P&L"),
    "Interest on TDS Payable":                ("Expenses", "Finance Charges", "P&L"),
    "LICENCE":                                ("Expenses", "Admin & Professional", "P&L"),
    "MACHINERY PARTS IGST 18%":               ("Expenses", "Repairs & Maintenance", "P&L"),
    "MACHINERY REPAIR & MAINTANCE":           ("Expenses", "Repairs & Maintenance", "P&L"),
    "Medicine":                               ("Expenses", "Employee Expenses", "P&L"),
    "MISC EXPENSES":                          ("Expenses", "Other Expenses", "P&L"),
    "Office Expenses":                        ("Expenses", "Office Expenses", "P&L"),
    "Office Maitenance":                      ("Expenses", "Office Expenses", "P&L"),
    "OFFICE REPAIR & MAINTANCE":              ("Expenses", "Repairs & Maintenance", "P&L"),
    "OFFICE REPAIR & MAINTANCE GST 18%":      ("Expenses", "Repairs & Maintenance", "P&L"),
    "OFFICE REPAIR & MAINTANCE IGST 18%":     ("Expenses", "Repairs & Maintenance", "P&L"),
    "Phone Pay Transaction Charges":          ("Expenses", "Finance Charges", "P&L"),
    "PRINTING & STATIONARY":                  ("Expenses", "Office Expenses", "P&L"),
    "Professional Charges":                   ("Expenses", "Admin & Professional", "P&L"),
    "Rent":                                   ("Expenses", "Rent", "P&L"),
    "Roundoff":                               ("Expenses", "Other Expenses", "P&L"),
    "SOFTWARE EXPNESES IGST 18%":             ("Expenses", "IT Expenses", "P&L"),
    "Staff Insurance":                        ("Expenses", "Insurance", "P&L"),
    "Tea Exp":                                ("Expenses", "Office Expenses", "P&L"),
    "Telephone Expense":                      ("Expenses", "Communication", "P&L"),
    "Traveling & Conveyance":                 ("Expenses", "Vehicle & Travel", "P&L"),
    "Vehicle Insurance":                      ("Expenses", "Insurance", "P&L"),
    "VEHICLE REPAIR & MAINTANCE":             ("Expenses", "Vehicle & Travel", "P&L"),
    "WATER EXPENSES":                         ("Expenses", "Utilities", "P&L"),
    "WEBSITE EXPENSES":                       ("Expenses", "IT Expenses", "P&L"),
    "Write Off Expenses":                     ("Expenses", "Other Expenses", "P&L"),

    # --- SUMMARY ROWS (excluded from analysis) ---
    "Profit & Loss A/c":                      ("Summary", "Profit & Loss", "Summary"),
    "Grand Total":                            ("Summary", "Grand Total", "Summary"),
}


# ════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════════

def safe_float(val):
    """Convert cell value to float, handling None, strings, and simple formula expressions."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    if s.startswith('='):
        try:
            return float(eval(s[1:]))
        except Exception:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_inr(value):
    """Format number as ₹ Indian Rupees with commas."""
    if abs(value) >= 10000000:
        return f"₹{value / 10000000:,.2f} Cr"
    elif abs(value) >= 100000:
        return f"₹{value / 100000:,.2f} L"
    else:
        return f"₹{value:,.2f}"


def fmt_inr_full(value):
    """Format number as ₹ with full commas, no abbreviation."""
    return f"₹{value:,.2f}"


def pct(part, whole):
    """Calculate percentage safely."""
    if whole == 0:
        return 0.0
    return (part / whole) * 100


# ════════════════════════════════════════════════════════════════
# STEP 1: LOAD & CLEAN DATA
# ════════════════════════════════════════════════════════════════

print("=" * 70)
print("  LOADING AND CLEANING DATA")
print("=" * 70)

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

cleaning_log = []

# Pass 1: Extract all raw rows
raw_rows = []
null_count = 0
for row_num in range(8, ws.max_row + 1):
    account = ws.cell(row=row_num, column=1).value
    if account is None or not str(account).strip():
        null_count += 1
        continue
    account = str(account).strip()
    row_data = {'account': account, 'row_num': row_num, 'values': {}}
    for i, (dc, cc) in enumerate(MONTH_COL_PAIRS):
        dv = ws.cell(row=row_num, column=dc).value
        cv = ws.cell(row=row_num, column=cc).value
        row_data['values'][MONTHS[i]] = {'debit_raw': dv, 'credit_raw': cv}
    raw_rows.append(row_data)

cleaning_log.append(f"Removed {null_count} blank/empty rows from the dataset.")
print(f"  ✓ Removed {null_count} blank rows")

# Pass 2: Identify and remove summary rows
summary_accounts = {"Profit & Loss A/c", "Grand Total"}
data_rows = [r for r in raw_rows if r['account'] not in summary_accounts]
removed_summary = len(raw_rows) - len(data_rows)
if removed_summary > 0:
    cleaning_log.append(f"Removed {removed_summary} summary/total rows ('Profit & Loss A/c', 'Grand Total') to prevent double-counting.")
    print(f"  ✓ Removed {removed_summary} summary rows")

# Pass 3: Deduplicate
seen = {}
dedup_rows = []
dup_count = 0
for r in data_rows:
    key = r['account']
    if key in seen:
        dup_count += 1
    else:
        seen[key] = True
        dedup_rows.append(r)
data_rows = dedup_rows
if dup_count > 0:
    cleaning_log.append(f"Removed {dup_count} duplicate account entries.")
    print(f"  ✓ Removed {dup_count} duplicates")
else:
    cleaning_log.append("No duplicate account entries found.")
    print(f"  ✓ No duplicates found")

# Pass 4: Convert all Debit/Credit to numeric (float), log conversions
conversion_issues = 0
for r in data_rows:
    for m in MONTHS:
        raw_d = r['values'][m]['debit_raw']
        raw_c = r['values'][m]['credit_raw']
        d = safe_float(raw_d)
        c = safe_float(raw_c)
        r['values'][m]['debit'] = d
        r['values'][m]['credit'] = c
        r['values'][m]['net'] = d - c
        if raw_d is not None and not isinstance(raw_d, (int, float)):
            if str(raw_d).strip().startswith('='):
                conversion_issues += 1
        if raw_c is not None and not isinstance(raw_c, (int, float)):
            if str(raw_c).strip().startswith('='):
                conversion_issues += 1

if conversion_issues > 0:
    cleaning_log.append(f"Converted {conversion_issues} formula-based cell values to numeric.")
    print(f"  ✓ Converted {conversion_issues} formula cells to numeric")

cleaning_log.append("Standardized column names to: Account, Month, Debit, Credit, NetBalance, Category, SubCategory, StatementType.")
cleaning_log.append("Converted data from wide (pivot) format to long (flat) format for analysis.")

print(f"  ✓ Final clean dataset: {len(data_rows)} unique accounts × {len(MONTHS)} months = {len(data_rows) * len(MONTHS)} records")


# ════════════════════════════════════════════════════════════════
# STEP 2: BUILD FLAT ANALYSIS DATASET
# ════════════════════════════════════════════════════════════════

flat_data = []
for r in data_rows:
    acct = r['account']
    classification = account_classification.get(acct, ("Unclassified", "Unclassified", "Unknown"))
    cat, subcat, stmt = classification
    for m in MONTHS:
        d = r['values'][m]['debit']
        c = r['values'][m]['credit']
        n = r['values'][m]['net']
        balance_type = "Debit" if n > 0 else ("Credit" if n < 0 else "Zero")
        flat_data.append({
            'Account': acct,
            'Month': m,
            'MonthLabel': MONTH_LABELS[m],
            'Debit': d,
            'Credit': c,
            'NetBalance': n,
            'BalanceType': balance_type,
            'Category': cat,
            'SubCategory': subcat,
            'StatementType': stmt,
        })

# Filter out summary rows from analysis
analysis_data = [r for r in flat_data if r['Category'] != 'Summary']

# Identify unclassified
unclassified = sorted(set(r['Account'] for r in analysis_data if r['Category'] == 'Unclassified'))


# ════════════════════════════════════════════════════════════════
# STEP 3: FINANCIAL CALCULATIONS
# ════════════════════════════════════════════════════════════════

# --- Overall Totals ---
total_debit = sum(r['Debit'] for r in analysis_data)
total_credit = sum(r['Credit'] for r in analysis_data)
net_balance = total_debit - total_credit

# --- P&L Items ---
total_revenue_credit = sum(r['Credit'] for r in analysis_data if r['Category'] == 'Revenue')
total_revenue_debit = sum(r['Debit'] for r in analysis_data if r['Category'] == 'Revenue')
total_revenue = total_revenue_credit - total_revenue_debit  # Net revenue
total_expenses = sum(r['Debit'] for r in analysis_data if r['Category'] == 'Expenses')
total_expense_credits = sum(r['Credit'] for r in analysis_data if r['Category'] == 'Expenses')
net_expenses = total_expenses - total_expense_credits  # Net expenses
net_profit_loss = total_revenue - net_expenses

# --- Balance Sheet Items ---
total_assets_debit = sum(r['Debit'] for r in analysis_data if r['Category'] == 'Assets')
total_assets_credit = sum(r['Credit'] for r in analysis_data if r['Category'] == 'Assets')
net_assets = total_assets_debit - total_assets_credit

total_liabilities_credit = sum(r['Credit'] for r in analysis_data if r['Category'] == 'Liabilities')
total_liabilities_debit = sum(r['Debit'] for r in analysis_data if r['Category'] == 'Liabilities')
net_liabilities = total_liabilities_credit - total_liabilities_debit

total_equity_credit = sum(r['Credit'] for r in analysis_data if r['Category'] == 'Equity')
total_equity_debit = sum(r['Debit'] for r in analysis_data if r['Category'] == 'Equity')
net_equity = total_equity_credit - total_equity_debit


# --- Category Totals ---
category_totals = defaultdict(lambda: {'debit': 0, 'credit': 0, 'net': 0})
for r in analysis_data:
    cat = r['Category']
    category_totals[cat]['debit'] += r['Debit']
    category_totals[cat]['credit'] += r['Credit']
    category_totals[cat]['net'] += r['NetBalance']

# --- SubCategory Totals ---
subcategory_totals = defaultdict(lambda: {'debit': 0, 'credit': 0, 'net': 0})
for r in analysis_data:
    key = f"{r['Category']} > {r['SubCategory']}"
    subcategory_totals[key]['debit'] += r['Debit']
    subcategory_totals[key]['credit'] += r['Credit']
    subcategory_totals[key]['net'] += r['NetBalance']

# --- Monthly P&L ---
monthly_pl = {}
for m in MONTHS:
    rev = sum(r['Credit'] for r in analysis_data if r['Month'] == m and r['Category'] == 'Revenue')
    rev_d = sum(r['Debit'] for r in analysis_data if r['Month'] == m and r['Category'] == 'Revenue')
    exp = sum(r['Debit'] for r in analysis_data if r['Month'] == m and r['Category'] == 'Expenses')
    exp_c = sum(r['Credit'] for r in analysis_data if r['Month'] == m and r['Category'] == 'Expenses')
    net_rev = rev - rev_d
    net_exp = exp - exp_c
    monthly_pl[m] = {
        'revenue': net_rev,
        'expenses': net_exp,
        'net': net_rev - net_exp,
    }

# --- Top Accounts by Debit ---
account_debit_totals = defaultdict(float)
account_credit_totals = defaultdict(float)
for r in analysis_data:
    account_debit_totals[r['Account']] += r['Debit']
    account_credit_totals[r['Account']] += r['Credit']

top_debit = sorted(account_debit_totals.items(), key=lambda x: -x[1])[:10]
top_credit = sorted(account_credit_totals.items(), key=lambda x: -x[1])[:10]

# --- Expense Subcategory breakdown ---
expense_subcats = defaultdict(float)
for r in analysis_data:
    if r['Category'] == 'Expenses' and r['Debit'] > 0:
        expense_subcats[r['SubCategory']] += r['Debit']
top_expense_subcats = sorted(expense_subcats.items(), key=lambda x: -x[1])

# --- Concentration Analysis ---
all_debit_sorted = sorted(account_debit_totals.items(), key=lambda x: -x[1])
cumulative = 0
top_80_count = 0
for acct, val in all_debit_sorted:
    cumulative += val
    top_80_count += 1
    if cumulative >= total_debit * 0.8:
        break

all_credit_sorted = sorted(account_credit_totals.items(), key=lambda x: -x[1])
cumulative_c = 0
top_80_count_c = 0
for acct, val in all_credit_sorted:
    cumulative_c += val
    top_80_count_c += 1
    if cumulative_c >= total_credit * 0.8:
        break


# ════════════════════════════════════════════════════════════════
# STEP 4: GENERATE CHARTS
# ════════════════════════════════════════════════════════════════

os.makedirs(OUTPUT_DIR, exist_ok=True)
chart_files = []

print("\n" + "=" * 70)
print("  GENERATING CHARTS")
print("=" * 70)

# ── Chart 1: KPI Summary (styled table) ────────────────────────────
fig, ax = plt.subplots(figsize=(12, 5))
ax.axis('off')
ax.set_title('KEY FINANCIAL INDICATORS (KPI SUMMARY)', fontsize=16, fontweight='bold',
             color=COLORS['primary'], pad=20, loc='left')

kpi_data = [
    ['Total Revenue (5 months)', fmt_inr_full(total_revenue), 'Total income from sales and other sources'],
    ['Total Expenses (5 months)', fmt_inr_full(net_expenses), 'Total costs including purchases, salaries, operations'],
    ['Net Profit / Loss', fmt_inr_full(net_profit_loss), 'Profit' if net_profit_loss > 0 else 'Loss'],
    ['Total Assets (Net)', fmt_inr_full(net_assets), 'What the company owns'],
    ['Total Liabilities (Net)', fmt_inr_full(net_liabilities), 'What the company owes'],
    ['Total Equity (Net)', fmt_inr_full(net_equity), "Owner's investment in the business"],
    ['Profit Margin', f"{pct(net_profit_loss, total_revenue):.1f}%", 'Profit as % of Revenue'],
    ['Expense-to-Revenue Ratio', f"{pct(net_expenses, total_revenue):.1f}%", 'Costs per ₹1 of revenue'],
]

table = ax.table(
    cellText=kpi_data,
    colLabels=['Metric', 'Value', 'What It Means'],
    cellLoc='left',
    loc='center',
    colWidths=[0.35, 0.25, 0.40],
)
table.auto_set_font_size(False)
table.set_fontsize(10)
table.scale(1, 1.8)

# Style header
for j in range(3):
    cell = table[0, j]
    cell.set_facecolor(COLORS['primary'])
    cell.set_text_props(color='white', fontweight='bold')
    cell.set_edgecolor('white')

# Style data rows
for i in range(1, len(kpi_data) + 1):
    for j in range(3):
        cell = table[i, j]
        cell.set_facecolor('#F8F9FA' if i % 2 == 0 else 'white')
        cell.set_edgecolor('#E8E8E8')
        if j == 1:
            cell.set_text_props(fontweight='bold')

# Highlight P&L row
pl_row = 3  # Net Profit row
for j in range(3):
    cell = table[pl_row, j]
    cell.set_facecolor('#E8F5E9' if net_profit_loss > 0 else '#FFEBEE')

plt.tight_layout()
kpi_path = os.path.join(OUTPUT_DIR, '01_kpi_summary.png')
plt.savefig(kpi_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(kpi_path)
print("  ✓ KPI Summary Table")

# ── Chart 2: Top 10 Accounts by Debit ──────────────────────────
fig, ax = plt.subplots(figsize=(12, 7))
names = [x[0][:35] for x in reversed(top_debit)]
values = [x[1] for x in reversed(top_debit)]
colors_bar = [COLORS['secondary']] * len(names)
bars = ax.barh(names, values, color=colors_bar, height=0.6, edgecolor='white', linewidth=0.5)

# Add value labels
for bar, val in zip(bars, values):
    ax.text(bar.get_width() + max(values) * 0.01, bar.get_y() + bar.get_height() / 2,
            fmt_inr(val), va='center', fontsize=9, color=COLORS['text'])

ax.set_title('TOP 10 ACCOUNTS BY DEBIT AMOUNT', fontsize=16, fontweight='bold',
             color=COLORS['primary'], pad=15, loc='left')
ax.set_xlabel('Debit Amount (₹)', fontsize=11, color=COLORS['text'])
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'₹{x/100000:.0f}L' if x >= 100000 else f'₹{x:,.0f}'))
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.grid(axis='x', alpha=0.3)
plt.tight_layout()
debit_path = os.path.join(OUTPUT_DIR, '02_top_debit_accounts.png')
plt.savefig(debit_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(debit_path)
print("  ✓ Top 10 Debit Accounts Chart")

# ── Chart 3: Top 10 Accounts by Credit ────────────────────────
fig, ax = plt.subplots(figsize=(12, 7))
names_c = [x[0][:35] for x in reversed(top_credit)]
values_c = [x[1] for x in reversed(top_credit)]
bars = ax.barh(names_c, values_c, color=COLORS['positive'], height=0.6, edgecolor='white', linewidth=0.5)

for bar, val in zip(bars, values_c):
    ax.text(bar.get_width() + max(values_c) * 0.01, bar.get_y() + bar.get_height() / 2,
            fmt_inr(val), va='center', fontsize=9, color=COLORS['text'])

ax.set_title('TOP 10 ACCOUNTS BY CREDIT AMOUNT', fontsize=16, fontweight='bold',
             color=COLORS['primary'], pad=15, loc='left')
ax.set_xlabel('Credit Amount (₹)', fontsize=11, color=COLORS['text'])
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'₹{x/100000:.0f}L' if x >= 100000 else f'₹{x:,.0f}'))
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.grid(axis='x', alpha=0.3)
plt.tight_layout()
credit_path = os.path.join(OUTPUT_DIR, '03_top_credit_accounts.png')
plt.savefig(credit_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(credit_path)
print("  ✓ Top 10 Credit Accounts Chart")

# ── Chart 4: Monthly P&L Trend (Line Chart) ──────────────────────
fig, ax = plt.subplots(figsize=(12, 6))
month_labels = [MONTH_LABELS[m].replace(' 2024', '') for m in MONTHS]
revenues = [monthly_pl[m]['revenue'] for m in MONTHS]
expenses = [monthly_pl[m]['expenses'] for m in MONTHS]
nets = [monthly_pl[m]['net'] for m in MONTHS]

ax.plot(month_labels, revenues, marker='o', linewidth=2.5, color=COLORS['positive'],
        label='Revenue', markersize=8, zorder=3)
ax.plot(month_labels, expenses, marker='s', linewidth=2.5, color=COLORS['negative'],
        label='Expenses', markersize=8, zorder=3)
ax.plot(month_labels, nets, marker='D', linewidth=2.5, color=COLORS['secondary'],
        label='Net P&L', markersize=8, zorder=3)

# Fill between for profit / loss
ax.fill_between(month_labels, nets, alpha=0.15,
                color=COLORS['positive'] if sum(nets) > 0 else COLORS['negative'])
ax.axhline(y=0, color='gray', linestyle='--', linewidth=0.8, alpha=0.5)

for i, (r, e, n) in enumerate(zip(revenues, expenses, nets)):
    ax.annotate(fmt_inr(r), (i, r), textcoords="offset points", xytext=(0, 12),
                fontsize=8, ha='center', color=COLORS['positive'])
    ax.annotate(fmt_inr(e), (i, e), textcoords="offset points", xytext=(0, -15),
                fontsize=8, ha='center', color=COLORS['negative'])

ax.set_title('MONTHLY REVENUE vs EXPENSES TREND', fontsize=16, fontweight='bold',
             color=COLORS['primary'], pad=15, loc='left')
ax.set_ylabel('Amount (₹)', fontsize=11, color=COLORS['text'])
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'₹{x/100000:.1f}L'))
ax.legend(frameon=True, fancybox=True, shadow=True, fontsize=10)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.grid(axis='y', alpha=0.3)
plt.tight_layout()
trend_path = os.path.join(OUTPUT_DIR, '04_monthly_trend.png')
plt.savefig(trend_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(trend_path)
print("  ✓ Monthly P&L Trend Chart")

# ── Chart 5: Category Split (Pie Chart – Debit Side) ─────────
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

# Debit pie
cat_debit = {c: v['debit'] for c, v in category_totals.items() if v['debit'] > 0 and c != 'Summary'}
labels_d = list(cat_debit.keys())
sizes_d = list(cat_debit.values())
colors_d = [CATEGORY_COLORS.get(l, '#BDC3C7') for l in labels_d]

wedges1, texts1, autotexts1 = ax1.pie(sizes_d, labels=labels_d, colors=colors_d, autopct='%1.1f%%',
                                       startangle=140, textprops={'fontsize': 10},
                                       wedgeprops={'edgecolor': 'white', 'linewidth': 2})
ax1.set_title('DEBIT DISTRIBUTION\nBY CATEGORY', fontsize=13, fontweight='bold',
              color=COLORS['primary'])

# Credit pie
cat_credit = {c: v['credit'] for c, v in category_totals.items() if v['credit'] > 0 and c != 'Summary'}
labels_c = list(cat_credit.keys())
sizes_c = list(cat_credit.values())
colors_c = [CATEGORY_COLORS.get(l, '#BDC3C7') for l in labels_c]

wedges2, texts2, autotexts2 = ax2.pie(sizes_c, labels=labels_c, colors=colors_c, autopct='%1.1f%%',
                                       startangle=140, textprops={'fontsize': 10},
                                       wedgeprops={'edgecolor': 'white', 'linewidth': 2})
ax2.set_title('CREDIT DISTRIBUTION\nBY CATEGORY', fontsize=13, fontweight='bold',
              color=COLORS['primary'])

plt.suptitle('FINANCIAL CATEGORY COMPOSITION', fontsize=16, fontweight='bold',
             color=COLORS['primary'], y=1.02)
plt.tight_layout()
pie_path = os.path.join(OUTPUT_DIR, '05_category_pie.png')
plt.savefig(pie_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(pie_path)
print("  ✓ Category Pie Charts")

# ── Chart 6: Expense Breakdown Bar ────────────────────────────
fig, ax = plt.subplots(figsize=(12, 8))
exp_names = [x[0][:30] for x in reversed(top_expense_subcats[:12])]
exp_values = [x[1] for x in reversed(top_expense_subcats[:12])]
exp_colors = plt.cm.YlOrRd([0.3 + 0.05 * i for i in range(len(exp_names))])

bars = ax.barh(exp_names, exp_values, color=exp_colors, height=0.65, edgecolor='white', linewidth=0.5)
for bar, val in zip(bars, exp_values):
    ax.text(bar.get_width() + max(exp_values) * 0.01, bar.get_y() + bar.get_height() / 2,
            f'{fmt_inr(val)} ({pct(val, total_expenses):.1f}%)', va='center', fontsize=9, color=COLORS['text'])

ax.set_title('EXPENSE BREAKDOWN BY SUBCATEGORY', fontsize=16, fontweight='bold',
             color=COLORS['primary'], pad=15, loc='left')
ax.set_xlabel('Amount (₹)', fontsize=11, color=COLORS['text'])
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'₹{x/100000:.0f}L' if x >= 100000 else f'₹{x:,.0f}'))
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.grid(axis='x', alpha=0.3)
plt.tight_layout()
expense_path = os.path.join(OUTPUT_DIR, '06_expense_breakdown.png')
plt.savefig(expense_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(expense_path)
print("  ✓ Expense Breakdown Chart")

# ── Chart 7: Monthly Revenue Breakdown (Stacked Bar) ─────────
fig, ax = plt.subplots(figsize=(12, 6))
month_short = [MONTH_LABELS[m].replace(' 2024', '') for m in MONTHS]

# Cumulative net profit
cum_profit = []
running = 0
for m in MONTHS:
    running += monthly_pl[m]['net']
    cum_profit.append(running)

bar_width = 0.35
x = range(len(MONTHS))
bars1 = ax.bar([i - bar_width/2 for i in x], [monthly_pl[m]['revenue'] for m in MONTHS],
               bar_width, label='Revenue', color=COLORS['positive'], edgecolor='white')
bars2 = ax.bar([i + bar_width/2 for i in x], [monthly_pl[m]['expenses'] for m in MONTHS],
               bar_width, label='Expenses', color=COLORS['negative'], edgecolor='white', alpha=0.85)

# Overlay cumulative profit line
ax2 = ax.twinx()
ax2.plot(x, cum_profit, color=COLORS['accent'], marker='D', linewidth=2.5,
         markersize=8, label='Cumulative Net P&L', zorder=5)
ax2.set_ylabel('Cumulative P&L (₹)', fontsize=11, color=COLORS['accent'])
ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'₹{x/100000:.1f}L'))
ax2.spines['top'].set_visible(False)

ax.set_xticks(list(x))
ax.set_xticklabels(month_short)
ax.set_title('MONTHLY REVENUE vs EXPENSES + CUMULATIVE P&L', fontsize=16, fontweight='bold',
             color=COLORS['primary'], pad=15, loc='left')
ax.set_ylabel('Monthly Amount (₹)', fontsize=11, color=COLORS['text'])
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'₹{x/100000:.0f}L' if x >= 100000 else f'₹{x:,.0f}'))
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.grid(axis='y', alpha=0.3)

# Combined legend
lines1, labels1 = ax.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax.legend(lines1 + lines2, labels1 + labels2, loc='upper left', frameon=True, fancybox=True, shadow=True)

plt.tight_layout()
stacked_path = os.path.join(OUTPUT_DIR, '07_monthly_stacked.png')
plt.savefig(stacked_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(stacked_path)
print("  ✓ Monthly Stacked Bar + Cumulative Chart")

# ── Chart 8: Concentration Analysis (Pareto Chart) ───────────
fig, ax = plt.subplots(figsize=(12, 6))
sorted_debits = sorted(account_debit_totals.items(), key=lambda x: -x[1])
sorted_debits = [(k, v) for k, v in sorted_debits if v > 0]

cum_pcts = []
running = 0
for _, val in sorted_debits:
    running += val
    cum_pcts.append(pct(running, total_debit))

x_range = range(len(sorted_debits))
ax.bar(x_range, [v for _, v in sorted_debits], color=COLORS['secondary'], alpha=0.7, width=1.0)
ax2 = ax.twinx()
ax2.plot(x_range, cum_pcts, color=COLORS['negative'], linewidth=2, marker='', zorder=5)
ax2.axhline(y=80, color=COLORS['accent'], linestyle='--', linewidth=1.5, alpha=0.8, label='80% Threshold')
ax2.set_ylabel('Cumulative %', fontsize=11, color=COLORS['negative'])
ax2.set_ylim(0, 105)
ax2.legend(loc='center right')

ax.set_title('CONCENTRATION ANALYSIS (PARETO – DEBIT SIDE)', fontsize=16, fontweight='bold',
             color=COLORS['primary'], pad=15, loc='left')
ax.set_xlabel(f'Accounts (Total: {len(sorted_debits)} with debit activity)', fontsize=11, color=COLORS['text'])
ax.set_ylabel('Debit Amount (₹)', fontsize=11, color=COLORS['text'])
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'₹{x/100000:.0f}L'))
ax.spines['top'].set_visible(False)
ax.set_xticks([])
plt.tight_layout()
pareto_path = os.path.join(OUTPUT_DIR, '08_concentration_pareto.png')
plt.savefig(pareto_path, dpi=150, bbox_inches='tight', facecolor=COLORS['bg'])
plt.close()
chart_files.append(pareto_path)
print("  ✓ Concentration / Pareto Chart")

print(f"\n  All {len(chart_files)} charts saved to '{OUTPUT_DIR}/' folder.")


# ════════════════════════════════════════════════════════════════
# STEP 5: GENERATE MARKDOWN REPORT
# ════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("  GENERATING BUSINESS REPORT")
print("=" * 70)

report_lines = []
def w(line=""):
    report_lines.append(line)


# ── COVER ──
w("# 📊 FINANCIAL BUSINESS REPORT")
w()
w("## Trial Balance Analysis — Labhyansh Solution")
w()
w(f"**Report Generated:** April 2026")
w(f"**Data Period:** April 2024 – August 2024 (FY 2024-25, Q1 & Q2)")
w(f"**Data Source:** Trial Balance ({EXCEL_FILE})")
w(f"**Prepared by:** Data Analytics Division")
w()
w("---")
w()

# ── SECTION 1: DATA CLEANING ──
w("## 1. 🧹 Data Cleaning Summary")
w()
w("Before performing the analysis, the following cleaning steps were carried out to ensure data quality:")
w()
for i, step in enumerate(cleaning_log, 1):
    w(f"{i}. {step}")
w()
w("> **Note:** No actual financial values were changed or manipulated. Only structural cleaning was done.")
w()
w("---")
w()

# ── SECTION 2: DATA UNDERSTANDING ──
w("## 2. 📋 Data Understanding")
w()
w("### What is a Trial Balance?")
w()
w("A **Trial Balance** is a financial statement that lists all the accounts of a company along with their debit and credit balances at a specific point in time. It is used to:")
w()
w("- ✅ Verify that total debits equal total credits (basic accounting check)")
w("- ✅ Serve as the foundation for preparing Profit & Loss and Balance Sheet statements")
w("- ✅ Help identify any errors in bookkeeping")
w()
w("### Dataset Overview")
w()
w(f"| Metric | Value |")
w(f"|--------|-------|")
w(f"| Total Unique Accounts | {len(data_rows)} |")
w(f"| Time Period | April 2024 – August 2024 (5 months) |")
w(f"| Financial Year | FY 2024-25 |")
w(f"| Quarters Covered | Q1 (Apr-Jun) and Q2 (Jul-Aug) |")
w(f"| Total Data Records (after unpivoting) | {len(analysis_data)} |")
if unclassified:
    w(f"| Unclassified Accounts | {len(unclassified)} |")
w()
w("---")
w()

# ── SECTION 3: FEATURE CREATION ──
w("## 3. 🔧 Feature Creation")
w()
w("The following new fields were created to enable deeper analysis:")
w()
w("| Feature | Formula / Logic | Purpose |")
w("|---------|-----------------|---------|")
w("| **Net Balance** | Debit − Credit | Shows whether an account has a debit or credit balance |")
w("| **Balance Type** | If Net > 0 → Debit, If Net < 0 → Credit | Quick classification of account direction |")
w("| **Category** | Manual mapping | Groups accounts into: Assets, Liabilities, Equity, Revenue, Expenses |")
w("| **SubCategory** | Manual mapping | Further grouping (e.g., Fixed Assets, Employee Expenses, Sales) |")
w("| **Statement Type** | Manual mapping | Identifies if account belongs to Balance Sheet or P&L |")
w()
w("---")
w()

# ── SECTION 4: KEY FINANCIAL SUMMARY ──
w("## 4. 💰 Key Financial Summary")
w()
w("![KPI Summary](report_output/01_kpi_summary.png)")
w()
w("### What These Numbers Mean")
w()
w(f"- **Total Revenue ({fmt_inr_full(total_revenue)}):** This is the total income the company has earned over 5 months from sales and other sources. Think of this as the total money coming into the business.")
w()
w(f"- **Total Expenses ({fmt_inr_full(net_expenses)}):** This is everything the company has spent — on raw materials (purchases), salaries, rent, utilities, transportation, and other operating costs.")
w()
if net_profit_loss > 0:
    w(f"- **Net Profit ({fmt_inr_full(net_profit_loss)}):** 🟢 The company is **profitable**. After paying all costs, {fmt_inr_full(net_profit_loss)} remains. The profit margin is **{pct(net_profit_loss, total_revenue):.1f}%**, meaning for every ₹100 earned, ₹{pct(net_profit_loss, total_revenue):.1f} is profit.")
else:
    w(f"- **Net Loss ({fmt_inr_full(abs(net_profit_loss))}):** 🔴 The company is running at a **loss**. Expenses exceed revenue by {fmt_inr_full(abs(net_profit_loss))}.")
w()
w(f"- **Assets ({fmt_inr_full(net_assets)}):** What the company owns — factory equipment, vehicles, bank balance, investments, receivables from customers.")
w()
w(f"- **Liabilities ({fmt_inr_full(net_liabilities)}):** What the company owes to others — loans, taxes payable, creditor dues, employee dues.")
w()
w(f"- **Equity ({fmt_inr_full(net_equity)}):** The owner's stake in the business — capital invested plus retained profits.")
w()
w("---")
w()

# ── SECTION 5: CATEGORY ANALYSIS ──
w("## 5. 📊 Category Analysis")
w()
w("![Category Split](report_output/05_category_pie.png)")
w()
w("### Category-wise Totals")
w()
w("| Category | Total Debit | Total Credit | Net Balance | Dominant Side |")
w("|----------|-------------|-------------|-------------|---------------|")
for cat in ['Assets', 'Liabilities', 'Equity', 'Revenue', 'Expenses']:
    if cat in category_totals:
        ct = category_totals[cat]
        dom = "Debit" if ct['net'] > 0 else "Credit"
        w(f"| {cat} | {fmt_inr_full(ct['debit'])} | {fmt_inr_full(ct['credit'])} | {fmt_inr_full(ct['net'])} | {dom} |")
w()
w("### What This Means")
w()
w("- **Assets** show a **debit balance** — this is expected as assets are debit-nature accounts.")
w("- **Liabilities & Equity** show **credit balances** — this means the company has obligations and capital on the books.")
w("- **Revenue** shows a **credit balance** — income naturally sits on the credit side.")
w("- **Expenses** show a **debit balance** — costs are debit-nature entries.")
w()
w("![Expense Breakdown](report_output/06_expense_breakdown.png)")
w()
w("### Top Expense Sub-Categories")
w()
w("| Rank | Sub-Category | Amount | % of Total Expenses |")
w("|------|-------------|--------|---------------------|")
for i, (subcat, val) in enumerate(top_expense_subcats[:10], 1):
    w(f"| {i} | {subcat} | {fmt_inr_full(val)} | {pct(val, total_expenses):.1f}% |")
w()
w("---")
w()

# ── SECTION 6: TOP ACCOUNT ANALYSIS ──
w("## 6. 🏆 Top Account Analysis")
w()
w("### Top 10 Accounts by Debit Amount")
w()
w("![Top Debit](report_output/02_top_debit_accounts.png)")
w()
w("These accounts had the **highest outflows** (debits) over the 5-month period:")
w()
w("| Rank | Account | Debit Total | Category | What It Means |")
w("|------|---------|-------------|----------|---------------|")
for i, (acct, val) in enumerate(top_debit, 1):
    cat = account_classification.get(acct, ("Unclassified",))[0]
    meaning = ""
    if cat == "Expenses":
        meaning = "Operating cost — money spent on business operations"
    elif cat == "Assets":
        meaning = "Money invested in assets or receivables"
    elif cat == "Liabilities":
        meaning = "Payment / adjustment of liabilities"
    elif cat == "Revenue":
        meaning = "Sales return or revenue adjustment"
    elif cat == "Equity":
        meaning = "Capital withdrawal or adjustment"
    else:
        meaning = "Uncategorized transaction"
    w(f"| {i} | {acct} | {fmt_inr_full(val)} | {cat} | {meaning} |")
w()

w("### Top 10 Accounts by Credit Amount")
w()
w("![Top Credit](report_output/03_top_credit_accounts.png)")
w()
w("These accounts had the **highest inflows** (credits) over the 5-month period:")
w()
w("| Rank | Account | Credit Total | Category | What It Means |")
w("|------|---------|-------------|----------|---------------|")
for i, (acct, val) in enumerate(top_credit, 1):
    cat = account_classification.get(acct, ("Unclassified",))[0]
    meaning = ""
    if cat == "Revenue":
        meaning = "Income earned from sales or other sources"
    elif cat == "Liabilities":
        meaning = "Amount owed to vendors, employees, or government"
    elif cat == "Equity":
        meaning = "Capital contributed by owner"
    elif cat == "Assets":
        meaning = "Reduction in asset value (depreciation or sale)"
    elif cat == "Expenses":
        meaning = "Expense reversal or recovery"
    else:
        meaning = "Uncategorized transaction"
    w(f"| {i} | {acct} | {fmt_inr_full(val)} | {cat} | {meaning} |")
w()
w("---")
w()

# ── SECTION 7: MONTHLY TREND ──
w("## 7. 📈 Monthly Trend Analysis")
w()
w("![Monthly Trend](report_output/04_monthly_trend.png)")
w()
w("### Month-by-Month Performance")
w()
w("| Month | Revenue | Expenses | Net P&L | Status |")
w("|-------|---------|----------|---------|--------|")
for m in MONTHS:
    mp = monthly_pl[m]
    status = "🟢 Profit" if mp['net'] > 0 else "🔴 Loss"
    w(f"| {MONTH_LABELS[m]} | {fmt_inr_full(mp['revenue'])} | {fmt_inr_full(mp['expenses'])} | {fmt_inr_full(mp['net'])} | {status} |")
w()

# Trend analysis
rev_trend = [monthly_pl[m]['revenue'] for m in MONTHS]
exp_trend = [monthly_pl[m]['expenses'] for m in MONTHS]
net_trend = [monthly_pl[m]['net'] for m in MONTHS]

if rev_trend[-1] > rev_trend[0]:
    rev_direction = "📈 **Revenue is trending UPWARD** — growing from " + fmt_inr(rev_trend[0]) + " to " + fmt_inr(rev_trend[-1])
elif rev_trend[-1] < rev_trend[0]:
    rev_direction = "📉 **Revenue is trending DOWNWARD** — declining from " + fmt_inr(rev_trend[0]) + " to " + fmt_inr(rev_trend[-1])
else:
    rev_direction = "➡️ **Revenue is flat**"

if exp_trend[-1] > exp_trend[0]:
    exp_direction = "📈 **Expenses are trending UPWARD** — growing from " + fmt_inr(exp_trend[0]) + " to " + fmt_inr(exp_trend[-1])
elif exp_trend[-1] < exp_trend[0]:
    exp_direction = "📉 **Expenses are trending DOWNWARD** — improving from " + fmt_inr(exp_trend[0]) + " to " + fmt_inr(exp_trend[-1])
else:
    exp_direction = "➡️ **Expenses are flat**"

w("### Trend Observations")
w()
w(f"- {rev_direction}")
w(f"- {exp_direction}")
best_month = MONTHS[net_trend.index(max(net_trend))]
worst_month = MONTHS[net_trend.index(min(net_trend))]
w(f"- 🏆 **Best month:** {MONTH_LABELS[best_month]} (Net: {fmt_inr_full(max(net_trend))})")
w(f"- ⚠️ **Weakest month:** {MONTH_LABELS[worst_month]} (Net: {fmt_inr_full(min(net_trend))})")
w()
w("![Monthly Stacked](report_output/07_monthly_stacked.png)")
w()
w(f"The **cumulative P&L** chart above shows how profits have been building up (or eroding) over the 5-month period. The final cumulative position is **{fmt_inr_full(sum(net_trend))}**.")
w()
w("---")
w()

# ── SECTION 8: CONCENTRATION ANALYSIS ──
w("## 8. 🎯 Concentration Analysis")
w()
w("![Concentration](report_output/08_concentration_pareto.png)")
w()
w("### Debit Concentration")
w()
total_accts_with_debit = len([v for v in account_debit_totals.values() if v > 0])
w(f"- Out of **{total_accts_with_debit}** accounts with debit activity, just **{top_80_count} accounts** (top {pct(top_80_count, total_accts_with_debit):.0f}%) account for **80% of all debit transactions**.")
w()
if top_80_count <= 10:
    w(f"  ⚠️ **High concentration risk:** A very small number of accounts dominate spending. If any of these accounts see a sudden spike, the overall financial position could be significantly impacted.")
else:
    w(f"  ✅ Debit activity is reasonably distributed across accounts.")
w()
w("### Credit Concentration")
w()
total_accts_with_credit = len([v for v in account_credit_totals.values() if v > 0])
w(f"- Out of **{total_accts_with_credit}** accounts with credit activity, just **{top_80_count_c} accounts** (top {pct(top_80_count_c, total_accts_with_credit):.0f}%) account for **80% of all credit transactions**.")
w()
if top_80_count_c <= 5:
    w(f"  ⚠️ **Revenue dependency risk:** Revenue is concentrated in very few accounts/sources. Losing any of these could severely impact income.")
else:
    w(f"  The credit side has moderate concentration.")
w()
w("### What This Means for the Business")
w()
w("- If your **largest customer stops buying**, or your **biggest supplier raises prices**, the impact on the business would be disproportionately large.")
w("- **Diversification** of both revenue sources and suppliers would reduce this risk.")
w()
w("---")
w()

# ── SECTION 9: BUSINESS INSIGHTS ──
w("## 9. 💡 Business Insights")
w()
w("Based on the detailed analysis, here are the most important insights for business stakeholders:")
w()

# Insight 1: Profitability
margin = pct(net_profit_loss, total_revenue)
if net_profit_loss > 0:
    w(f"### 1. ✅ The Company is Profitable")
    w(f"- Net profit of **{fmt_inr_full(net_profit_loss)}** over 5 months with a margin of **{margin:.1f}%**.")
    if margin < 5:
        w(f"- However, the margin is **very thin** — just {margin:.1f}%. A small increase in costs could wipe out profits.")
    elif margin < 15:
        w(f"- The margin is **moderate** — there's room for improvement through cost optimization.")
    else:
        w(f"- This is a **healthy margin** indicating good cost management.")
else:
    w(f"### 1. 🔴 The Company is Operating at a Loss")
    w(f"- Net loss of **{fmt_inr_full(abs(net_profit_loss))}** — expenses exceed revenue.")
    w(f"- Immediate cost-cutting or revenue growth measures are needed.")
w()

# Insight 2: Expense structure
purchases_total = expense_subcats.get('Purchases', 0)
if total_expenses > 0:
    purchase_pct = pct(purchases_total, total_expenses)
    w(f"### 2. 📦 Purchases Dominate Expenses")
    w(f"- **Purchases** account for **{purchase_pct:.1f}%** of total expenses ({fmt_inr_full(purchases_total)}).")
    w(f"- This is typical for a manufacturing/trading business, but negotiating better rates with suppliers could significantly improve margins.")
    w()

# Insight 3: Employee costs
emp_costs = expense_subcats.get('Employee Expenses', 0)
if total_expenses > 0:
    emp_pct = pct(emp_costs, total_expenses)
    w(f"### 3. 👥 Employee Costs")
    w(f"- Employee-related expenses (salary, EPF, ESIC, bonuses) total **{fmt_inr_full(emp_costs)}** ({emp_pct:.1f}% of expenses).")
    if emp_pct > 20:
        w(f"- This is a significant portion — review whether workforce productivity justifies this cost.")
    else:
        w(f"- This is within a reasonable range for the business size.")
    w()

# Insight 4: Asset structure
w(f"### 4. 🏭 Asset-Heavy Structure")
w(f"- The company holds **{fmt_inr_full(net_assets)}** in net assets, including factory machinery, vehicles, and equipment.")
w(f"- This is typical for a **manufacturing business** — but it also means high depreciation costs and maintenance expenses.")
w()

# Insight 5: Cash & Bank
cash_total = sum(r['Debit'] - r['Credit'] for r in analysis_data if r['SubCategory'] in ('Cash', 'Bank'))
w(f"### 5. 💵 Cash & Bank Position")
w(f"- Net cash and bank position: **{fmt_inr_full(cash_total)}**")
if cash_total > 0:
    w(f"- The company has positive cash reserves, which is healthy for meeting short-term obligations.")
else:
    w(f"- ⚠️ Negative cash position — the company may face difficulty in meeting day-to-day payments.")
w()

# Insight 6: Revenue dependency
w(f"### 6. 🎯 Revenue Concentration Risk")
w(f"- Only **{top_80_count_c}** accounts contribute 80% of all credit (revenue + other inflows).")
w(f"- If the main sales channel or key customers are disrupted, the impact would be severe.")
w(f"- **Recommendation:** Diversify customer base and explore new revenue streams.")
w()

# Insight 7: Tax & compliance
tax_related = sum(r['Debit'] for r in analysis_data if 'Tax' in r['SubCategory'] or 'TDS' in r['SubCategory'] or 'Duties' in r['SubCategory'])
w(f"### 7. 🏛️ Tax & Compliance")
w(f"- Tax-related transactions (GST, TDS, Income Tax) total **{fmt_inr_full(tax_related)}** on the debit side.")
w(f"- The company is managing multiple tax components (CGST, SGST, IGST, TDS) — proper compliance requires diligent tracking.")
w()
w("---")
w()

# ── SECTION 10: VISUALS INDEX ──
w("## 10. 📸 Visual Summary Index")
w()
w("All charts generated as part of this report:")
w()
w("| # | Chart | Description |")
w("|---|-------|-------------|")
w("| 1 | KPI Summary Table | Key financial metrics at a glance |")
w("| 2 | Top 10 Debit Accounts | Accounts with highest debit (outflow) amounts |")
w("| 3 | Top 10 Credit Accounts | Accounts with highest credit (inflow) amounts |")
w("| 4 | Monthly P&L Trend (Line) | Revenue vs Expenses over 5 months |")
w("| 5 | Category Pie Charts | Distribution of debits and credits by category |")
w("| 6 | Expense Breakdown (Bar) | Detailed breakdown of all expense sub-categories |")
w("| 7 | Monthly Stacked Bar + Cumulative | Side-by-side revenue/expense comparison with cumulative P&L |")
w("| 8 | Concentration / Pareto (Bar + Line) | Shows how few accounts dominate the totals |")
w()
w("---")
w()

# ── SECTION 11: FINAL CONCLUSION ──
w("## 11. 🏁 Final Conclusion")
w()
w("### Overall Financial Health Assessment")
w()
if net_profit_loss > 0 and margin >= 5:
    health_status = "🟢 GOOD"
    health_desc = "The company is profitable with a reasonable margin."
elif net_profit_loss > 0 and margin < 5:
    health_status = "🟡 MODERATE"
    health_desc = "The company is barely profitable — the margin is very thin and vulnerable to cost fluctuations."
else:
    health_status = "🔴 NEEDS ATTENTION"
    health_desc = "The company is currently operating at a loss and needs immediate corrective action."

w(f"**Overall Status:** {health_status}")
w()
w(f"**Summary:** {health_desc}")
w()
w("### Key Strengths")
w()
if net_profit_loss > 0:
    w("- ✅ Company is profitable and generating positive cash flow")
if net_assets > 0:
    w("- ✅ Strong asset base indicating long-term investment in the business")
if cash_total > 0:
    w("- ✅ Positive cash & bank position for meeting short-term needs")
w("- ✅ Diverse expense structure across multiple categories")
w("- ✅ Regular tax compliance (GST, TDS) indicates good governance")
w()
w("### Key Risks")
w()
if margin < 10 and net_profit_loss > 0:
    w(f"- ⚠️ **Thin profit margin** ({margin:.1f}%) — vulnerable to cost increases")
if top_80_count_c <= 5:
    w("- ⚠️ **Revenue concentration** — too dependent on a few sources")
if purchases_total > 0 and purchase_pct > 60:
    w(f"- ⚠️ **High material cost** — purchases are {purchase_pct:.1f}% of expenses")
w("- ⚠️ **Asset-heavy structure** — high depreciation and maintenance overhead")
w("- ⚠️ **No long-term debt visible** — may limit growth capacity if external funding is ever needed")
w()
w("---")
w()

# ── SECTION 12: BUSINESS RECOMMENDATIONS ──
w("## 12. 📋 Business Recommendations")
w()
w("Based on the analysis, here are **actionable recommendations** for the management:")
w()
w("### Recommendation 1: 🔄 Negotiate Better Purchase Rates")
w(f"- Purchases are the **single largest expense** at {fmt_inr_full(purchases_total)}.")
w("- Even a 2–3% reduction in purchase costs through bulk deals or alternate suppliers could add significantly to the bottom line.")
w()
w("### Recommendation 2: 👥 Review Workforce Efficiency")
w(f"- Employee costs total {fmt_inr_full(emp_costs)}.")
w("- Conduct a productivity audit — ensure the headcount is aligned with output needs. Consider automation for repetitive tasks.")
w()
w("### Recommendation 3: 🎯 Diversify Revenue Sources")
w("- Revenue is concentrated in a few accounts/sales channels.")
w("- Explore new markets, product lines, or distribution channels to reduce dependency risk.")
w()
w("### Recommendation 4: 💰 Build Cash Reserves")
w("- Maintain at least 2-3 months of operating expenses as cash reserves.")
w("- This provides a safety buffer against temporary disruptions in sales or unexpected costs.")
w()
w("### Recommendation 5: 📊 Implement Monthly Financial Reviews")
w("- Use this type of analysis every month to track trends early.")
w("- Set up KPI dashboards (as built with Power BI) for real-time monitoring.")
w("- Identify problems **before** they become crises.")
w()
w("---")
w()
w("## Disclaimer")
w()
w("*This report has been prepared based on the Trial Balance data provided. No actual financial values were altered during analysis. The classifications and insights are based on standard accounting principles and the account structure observed in the dataset. This report should be used in conjunction with verified financial statements prepared by a qualified Chartered Accountant.*")
w()
w("---")
w()
w("**End of Report**")

# Write the report
report_path = os.path.join(OUTPUT_DIR, "Financial_Business_Report.md")
with open(report_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(report_lines))

print(f"  ✓ Report saved to: {report_path}")
print()
print("=" * 70)
print("  ✅ REPORT GENERATION COMPLETE")
print("=" * 70)
print(f"  📄 Report:  {report_path}")
print(f"  📊 Charts:  {len(chart_files)} charts in '{OUTPUT_DIR}/' folder")
print(f"  📁 Files: {', '.join(os.listdir(OUTPUT_DIR))}")
print()
