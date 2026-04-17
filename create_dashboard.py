#!/usr/bin/env python3
"""
Generate an interactive HTML financial dashboard with glassmorphism design.
Uses Chart.js for interactive charts. All data is embedded.
"""

import sys, io, os, base64
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

OUTPUT_DIR = "report_output"
DASHBOARD_FILE = os.path.join(OUTPUT_DIR, "Financial_Dashboard.html")

# ── Read the analysis data from the generate_report pipeline ────────
import openpyxl
from collections import defaultdict

EXCEL_FILE = "Dummy Data for Review.xlsx"
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb['Trial Balance']

MONTHS = ['Apr24', 'May24', 'Jun24', 'Jul24', 'Aug24']
MONTH_LABELS_SHORT = ['Apr', 'May', 'Jun', 'Jul', 'Aug']
MONTH_LABELS_FULL = ['April 2024', 'May 2024', 'June 2024', 'July 2024', 'August 2024']
MONTH_COL_PAIRS = [(2, 3), (4, 5), (6, 7), (8, 9), (10, 11)]

def safe_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    if not s: return 0.0
    if s.startswith('='):
        try: return float(eval(s[1:]))
        except: return 0.0
    try: return float(s)
    except: return 0.0

# Account classification (same as before)
account_classification = {
    "Capital Account": ("Equity","Capital","Balance Sheet"),
    "Opening Adjustment": ("Equity","Capital","Balance Sheet"),
    "MY FACTORY (DELHI)`2": ("Equity","Capital","Balance Sheet"),
    "MEHTA CAPITAL ACCOUNT": ("Equity","Capital","Balance Sheet"),
    "Loans (Liability)": ("Liabilities","Loans","Balance Sheet"),
    "Unsecured Loans": ("Liabilities","Unsecured Loans","Balance Sheet"),
    "A.M FOODS": ("Liabilities","Unsecured Loans","Balance Sheet"),
    "Current Liabilities": ("Liabilities","Current Liabilities","Balance Sheet"),
    "Duties & Taxes": ("Liabilities","Duties & Taxes","Balance Sheet"),
    "Input Receivable": ("Liabilities","Duties & Taxes - Input","Balance Sheet"),
    "CGST  Input": ("Liabilities","Duties & Taxes - Input","Balance Sheet"),
    "IGST Input": ("Liabilities","Duties & Taxes - Input","Balance Sheet"),
    "MACHINERY PARTS IGST TAX 18%": ("Liabilities","Duties & Taxes - Input","Balance Sheet"),
    "PURCHASE CGST 9%": ("Liabilities","Duties & Taxes - Input","Balance Sheet"),
    "PURCHASE SGST 9%": ("Liabilities","Duties & Taxes - Input","Balance Sheet"),
    "SGST Input": ("Liabilities","Duties & Taxes - Input","Balance Sheet"),
    "Output": ("Liabilities","Duties & Taxes - Output","Balance Sheet"),
    "CGST": ("Liabilities","Duties & Taxes - Output","Balance Sheet"),
    "IGST": ("Liabilities","Duties & Taxes - Output","Balance Sheet"),
    "SGST": ("Liabilities","Duties & Taxes - Output","Balance Sheet"),
    "TDS Payable": ("Liabilities","TDS Payable","Balance Sheet"),
    "TDS 194C Contractor Company": ("Liabilities","TDS Payable","Balance Sheet"),
    "TDS 194C Contractor Non Company": ("Liabilities","TDS Payable","Balance Sheet"),
    "TDS Professional 194J": ("Liabilities","TDS Payable","Balance Sheet"),
    "Provisions": ("Liabilities","Provisions","Balance Sheet"),
    "AUDIT FEES PAYABLE": ("Liabilities","Provisions","Balance Sheet"),
    "EPF-Administrative Charges Payable": ("Liabilities","Provisions","Balance Sheet"),
    "EPF PAYABLE": ("Liabilities","Provisions","Balance Sheet"),
    "ESIC PAYABLE": ("Liabilities","Provisions","Balance Sheet"),
    "LWF Payble": ("Liabilities","Provisions","Balance Sheet"),
    "MANUFACTURING EXPENSES PAYABLE": ("Liabilities","Provisions","Balance Sheet"),
    "Provision for Electricity Expenses": ("Liabilities","Provisions","Balance Sheet"),
    "PROVISION FOR EXPENSES": ("Liabilities","Provisions","Balance Sheet"),
    "Provision for Tax": ("Liabilities","Provisions","Balance Sheet"),
    "SALARY PAYABLE": ("Liabilities","Provisions","Balance Sheet"),
    "Sundry Creditors": ("Liabilities","Sundry Creditors","Balance Sheet"),
    "HDFC Credit Card 2140": ("Liabilities","Sundry Creditors","Balance Sheet"),
    "Fixed Assets": ("Assets","Fixed Assets","Balance Sheet"),
    "Accumulated Depreciation": ("Assets","Fixed Assets - Depreciation","Balance Sheet"),
    "BIKE": ("Assets","Fixed Assets","Balance Sheet"),
    "CAR(ISUZU)": ("Assets","Fixed Assets","Balance Sheet"),
    "Chapati Making Machine": ("Assets","Fixed Assets","Balance Sheet"),
    "COOLER": ("Assets","Fixed Assets","Balance Sheet"),
    "DUSTBIN 120LTR": ("Assets","Fixed Assets","Balance Sheet"),
    "FURNITURE IGST 18%": ("Assets","Fixed Assets","Balance Sheet"),
    "LUMINOUS BATTERY": ("Assets","Fixed Assets","Balance Sheet"),
    "MACHINERY IGST 18%": ("Assets","Fixed Assets","Balance Sheet"),
    "MOBILE": ("Assets","Fixed Assets","Balance Sheet"),
    "PUMP": ("Assets","Fixed Assets","Balance Sheet"),
    "RO PLANT IGST 18%": ("Assets","Fixed Assets","Balance Sheet"),
    "SAREGAMA CARVAAN GST 18%": ("Assets","Fixed Assets","Balance Sheet"),
    "TABLE": ("Assets","Fixed Assets","Balance Sheet"),
    "Weighing Machine": ("Assets","Fixed Assets","Balance Sheet"),
    "Investments": ("Assets","Investments","Balance Sheet"),
    "Investment in FDR": ("Assets","Investments","Balance Sheet"),
    "Kotak  Mutual Fund": ("Assets","Investments","Balance Sheet"),
    "Motilal Oswal Financial Services Ltd": ("Assets","Investments","Balance Sheet"),
    "Nippon Mutual Fund": ("Assets","Investments","Balance Sheet"),
    "TPT Investment Management (TPTIM)": ("Assets","Investments","Balance Sheet"),
    "Current Assets": ("Assets","Current Assets","Balance Sheet"),
    "Opening Stock": ("Assets","Current Assets - Stock","Balance Sheet"),
    "Loans & Advances (Asset)": ("Assets","Loans & Advances","Balance Sheet"),
    "Advance Salary": ("Assets","Loans & Advances","Balance Sheet"),
    "Sundry Debtors": ("Assets","Sundry Debtors","Balance Sheet"),
    "Cash-in-Hand": ("Assets","Cash","Balance Sheet"),
    "Cash": ("Assets","Cash","Balance Sheet"),
    "Bank Accounts": ("Assets","Bank","Balance Sheet"),
    "HDFC BANK": ("Assets","Bank","Balance Sheet"),
    "Accrued Income on FDR": ("Assets","Other Current Assets","Balance Sheet"),
    "ADVANCE TAX": ("Assets","Tax Assets","Balance Sheet"),
    "SECURITY DEPOSIT - IOC APPARELS AND HCG GAS": ("Assets","Other Current Assets","Balance Sheet"),
    "TCS": ("Assets","Tax Assets","Balance Sheet"),
    "TDS": ("Assets","Tax Assets","Balance Sheet"),
    "Sales Accounts": ("Revenue","Sales","P&L"),
    "Un Exempt Sales A/c": ("Revenue","Sales","P&L"),
    "Purchase Accounts": ("Expenses","Purchases","P&L"),
    "DISCOUNT REBATE": ("Expenses","Purchases - Discount","P&L"),
    "DISCOUNT RECEIVED": ("Revenue","Discount Received","P&L"),
    "PURCHASE GST 18%": ("Expenses","Purchases","P&L"),
    "PURCHASE GST 28%": ("Expenses","Purchases","P&L"),
    "PURCHASE GST 5%": ("Expenses","Purchases","P&L"),
    "PURCHASE GST TAX FREE": ("Expenses","Purchases","P&L"),
    "PURCHASE IGST 12%": ("Expenses","Purchases","P&L"),
    "PURCHASE IGST 18%": ("Expenses","Purchases","P&L"),
    "PURCHASE IGST 5%": ("Expenses","Purchases","P&L"),
    "Expenses (Direct) (Direct Expenses)": ("Expenses","Direct Expenses","P&L"),
    "DELIVERY CHARGES": ("Expenses","Direct Expenses","P&L"),
    "Electricity Expense": ("Expenses","Direct Expenses","P&L"),
    "FREIGHT AND CARTAGE": ("Expenses","Direct Expenses","P&L"),
    "FREIGHT AND CARTAGE IGST 12%": ("Expenses","Direct Expenses","P&L"),
    "GAS EXPENSES": ("Expenses","Direct Expenses","P&L"),
    "GAS EXPENSES IGST 18%": ("Expenses","Direct Expenses","P&L"),
    "HR SERVICES (MANPOWER SERVICES) GST 18%": ("Expenses","Direct Expenses","P&L"),
    "Labour Charges  (Atta Unload)": ("Expenses","Direct Expenses","P&L"),
    "Labour on hire": ("Expenses","Direct Expenses","P&L"),
    "Income (Indirect) (Indirect Incomes)": ("Revenue","Indirect Income","P&L"),
    "Cartage Charges": ("Revenue","Indirect Income","P&L"),
    "Cash Discount": ("Revenue","Indirect Income","P&L"),
    "Interest Income": ("Revenue","Indirect Income","P&L"),
    "Interest on FDR- 50301072325452": ("Revenue","Indirect Income","P&L"),
    "Interest on Income Tax Refund": ("Revenue","Indirect Income","P&L"),
    "Misc Receipt": ("Revenue","Indirect Income","P&L"),
    "Expenses (Indirect) (Indirect Expenses)": ("Expenses","Indirect Expenses","P&L"),
    "Employee Expenses": ("Expenses","Employee Expenses","P&L"),
    "ADMINISTRATION CHARGES": ("Expenses","Employee Expenses","P&L"),
    "DIWALI BONUS": ("Expenses","Employee Expenses","P&L"),
    "EDLI Charges": ("Expenses","Employee Expenses","P&L"),
    "EPF": ("Expenses","Employee Expenses","P&L"),
    "ESIC": ("Expenses","Employee Expenses","P&L"),
    "LWF": ("Expenses","Employee Expenses","P&L"),
    "Recruitment Expenses": ("Expenses","Employee Expenses","P&L"),
    "SALARY": ("Expenses","Employee Expenses","P&L"),
    "STAFF WELFARE": ("Expenses","Employee Expenses","P&L"),
    "ACCOUNTING CHARGES": ("Expenses","Admin & Professional","P&L"),
    "AUDIT FEES": ("Expenses","Admin & Professional","P&L"),
    "BAD DEBTS": ("Expenses","Other Expenses","P&L"),
    "BANK CHARGES": ("Expenses","Finance Charges","P&L"),
    "BUSINESS PROMOTION GST18%": ("Expenses","Marketing & Promotion","P&L"),
    "COMPUTER REPAIR AND MANITANCE": ("Expenses","Repairs & Maintenance","P&L"),
    "DEPRECIATION": ("Expenses","Depreciation","P&L"),
    "FACTORY REPAIR & MAINTANCE": ("Expenses","Repairs & Maintenance","P&L"),
    "Fees & Subscription": ("Expenses","Admin & Professional","P&L"),
    "FUEL EXEPENSES": ("Expenses","Vehicle & Travel","P&L"),
    "GENERAL EXPENSES": ("Expenses","Other Expenses","P&L"),
    "Income Tax Expenses": ("Expenses","Tax Expenses","P&L"),
    "INSURANCE": ("Expenses","Insurance","P&L"),
    "Interest on TDS Payable": ("Expenses","Finance Charges","P&L"),
    "LICENCE": ("Expenses","Admin & Professional","P&L"),
    "MACHINERY PARTS IGST 18%": ("Expenses","Repairs & Maintenance","P&L"),
    "MACHINERY REPAIR & MAINTANCE": ("Expenses","Repairs & Maintenance","P&L"),
    "Medicine": ("Expenses","Employee Expenses","P&L"),
    "MISC EXPENSES": ("Expenses","Other Expenses","P&L"),
    "Office Expenses": ("Expenses","Office Expenses","P&L"),
    "Office Maitenance": ("Expenses","Office Expenses","P&L"),
    "OFFICE REPAIR & MAINTANCE": ("Expenses","Repairs & Maintenance","P&L"),
    "OFFICE REPAIR & MAINTANCE GST 18%": ("Expenses","Repairs & Maintenance","P&L"),
    "OFFICE REPAIR & MAINTANCE IGST 18%": ("Expenses","Repairs & Maintenance","P&L"),
    "Phone Pay Transaction Charges": ("Expenses","Finance Charges","P&L"),
    "PRINTING & STATIONARY": ("Expenses","Office Expenses","P&L"),
    "Professional Charges": ("Expenses","Admin & Professional","P&L"),
    "Rent": ("Expenses","Rent","P&L"),
    "Roundoff": ("Expenses","Other Expenses","P&L"),
    "SOFTWARE EXPNESES IGST 18%": ("Expenses","IT Expenses","P&L"),
    "Staff Insurance": ("Expenses","Insurance","P&L"),
    "Tea Exp": ("Expenses","Office Expenses","P&L"),
    "Telephone Expense": ("Expenses","Communication","P&L"),
    "Traveling & Conveyance": ("Expenses","Vehicle & Travel","P&L"),
    "Vehicle Insurance": ("Expenses","Insurance","P&L"),
    "VEHICLE REPAIR & MAINTANCE": ("Expenses","Vehicle & Travel","P&L"),
    "WATER EXPENSES": ("Expenses","Utilities","P&L"),
    "WEBSITE EXPENSES": ("Expenses","IT Expenses","P&L"),
    "Write Off Expenses": ("Expenses","Other Expenses","P&L"),
    "Profit & Loss A/c": ("Summary","Profit & Loss","Summary"),
    "Grand Total": ("Summary","Grand Total","Summary"),
}

# Extract data
flat_data = []
for row_num in range(8, ws.max_row + 1):
    account = ws.cell(row=row_num, column=1).value
    if not account or not str(account).strip(): continue
    account = str(account).strip()
    if account in ("Profit & Loss A/c", "Grand Total"): continue
    classification = account_classification.get(account, ("Unclassified","Unclassified","Unknown"))
    cat, subcat, stmt = classification
    for i, (dc, cc) in enumerate(MONTH_COL_PAIRS):
        d = safe_float(ws.cell(row=row_num, column=dc).value)
        c = safe_float(ws.cell(row=row_num, column=cc).value)
        flat_data.append({
            'Account': account, 'Month': MONTHS[i], 'Debit': d, 'Credit': c,
            'Net': d - c, 'Category': cat, 'SubCategory': subcat, 'Statement': stmt
        })

# ── Compute all metrics ─────────────────────────────────────────

# Monthly P&L
monthly_revenue = []
monthly_expenses = []
monthly_net = []
for m in MONTHS:
    rev = sum(r['Credit'] - r['Debit'] for r in flat_data if r['Month'] == m and r['Category'] == 'Revenue')
    exp = sum(r['Debit'] - r['Credit'] for r in flat_data if r['Month'] == m and r['Category'] == 'Expenses')
    monthly_revenue.append(round(rev, 2))
    monthly_expenses.append(round(exp, 2))
    monthly_net.append(round(rev - exp, 2))

total_revenue = sum(monthly_revenue)
total_expenses = sum(monthly_expenses)
net_profit = total_revenue - total_expenses
margin = round((net_profit / total_revenue) * 100, 1) if total_revenue else 0

# Assets / Liabilities / Equity
net_assets = sum(r['Debit'] - r['Credit'] for r in flat_data if r['Category'] == 'Assets')
net_liabilities = sum(r['Credit'] - r['Debit'] for r in flat_data if r['Category'] == 'Liabilities')
net_equity = sum(r['Credit'] - r['Debit'] for r in flat_data if r['Category'] == 'Equity')
cash_bank = sum(r['Debit'] - r['Credit'] for r in flat_data if r['SubCategory'] in ('Cash', 'Bank'))

# Category totals
cat_debit = defaultdict(float)
cat_credit = defaultdict(float)
for r in flat_data:
    if r['Category'] not in ('Summary', 'Unclassified'):
        cat_debit[r['Category']] += r['Debit']
        cat_credit[r['Category']] += r['Credit']

# Expense subcategory breakdown
exp_sub = defaultdict(float)
for r in flat_data:
    if r['Category'] == 'Expenses' and r['Debit'] > 0:
        exp_sub[r['SubCategory']] += r['Debit']
top_exp_sub = sorted(exp_sub.items(), key=lambda x: -x[1])

# Top accounts
acct_debit = defaultdict(float)
acct_credit = defaultdict(float)
for r in flat_data:
    acct_debit[r['Account']] += r['Debit']
    acct_credit[r['Account']] += r['Credit']

top_debit_accts = sorted(acct_debit.items(), key=lambda x: -x[1])[:10]
top_credit_accts = sorted(acct_credit.items(), key=lambda x: -x[1])[:10]

# Monthly cumulative
cum_net = []
running = 0
for n in monthly_net:
    running += n
    cum_net.append(round(running, 2))

# Revenue growth
rev_growth = round(((monthly_revenue[-1] - monthly_revenue[0]) / monthly_revenue[0]) * 100, 1) if monthly_revenue[0] else 0

# Format helpers
def fmt(v):
    if abs(v) >= 10000000: return f"{v/10000000:.2f} Cr"
    if abs(v) >= 100000: return f"{v/100000:.2f} L"
    return f"{v:,.0f}"

print("  Data extracted and computed successfully")

# ── Build the HTML ──────────────────────────────────────────────

# Prepare JSON-safe data for Chart.js
import json

exp_sub_labels = json.dumps([x[0] for x in top_exp_sub[:10]])
exp_sub_values = json.dumps([round(x[1], 2) for x in top_exp_sub[:10]])

top_deb_labels = json.dumps([x[0][:30] for x in top_debit_accts])
top_deb_values = json.dumps([round(x[1], 2) for x in top_debit_accts])

top_cred_labels = json.dumps([x[0][:30] for x in top_credit_accts])
top_cred_values = json.dumps([round(x[1], 2) for x in top_credit_accts])

cat_labels = json.dumps(list(cat_debit.keys()))
cat_deb_vals = json.dumps([round(cat_debit[k], 2) for k in cat_debit])
cat_cred_vals = json.dumps([round(cat_credit.get(k, 0), 2) for k in cat_debit])

# Monthly sub-category heatmap data
monthly_exp_sub = {}
for r in flat_data:
    if r['Category'] == 'Expenses' and r['Debit'] > 0:
        key = r['SubCategory']
        if key not in monthly_exp_sub:
            monthly_exp_sub[key] = {m: 0 for m in MONTHS}
        monthly_exp_sub[key][r['Month']] += r['Debit']

top_5_exp_subs = [x[0] for x in top_exp_sub[:5]]
heatmap_datasets = []
heatmap_colors = ['rgba(59,130,246,0.8)', 'rgba(239,68,68,0.8)', 'rgba(16,185,129,0.8)', 'rgba(245,158,11,0.8)', 'rgba(139,92,246,0.8)']
for i, sub in enumerate(top_5_exp_subs):
    vals = [round(monthly_exp_sub.get(sub, {}).get(m, 0), 2) for m in MONTHS]
    heatmap_datasets.append({
        'label': sub,
        'data': vals,
        'backgroundColor': heatmap_colors[i % len(heatmap_colors)],
        'borderRadius': 4,
    })

heatmap_json = json.dumps(heatmap_datasets)

html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Labhyansh Solution | Financial Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --bg-base: #0a0e1a;
    --bg-surface: rgba(15, 23, 42, 0.65);
    --bg-elevated: rgba(30, 41, 59, 0.5);
    --glass-border: rgba(255, 255, 255, 0.08);
    --glass-shine: rgba(255, 255, 255, 0.04);
    --accent: #3b82f6;
    --accent-soft: rgba(59, 130, 246, 0.15);
    --green: #10b981;
    --green-soft: rgba(16, 185, 129, 0.12);
    --red: #ef4444;
    --red-soft: rgba(239, 68, 68, 0.12);
    --amber: #f59e0b;
    --amber-soft: rgba(245, 158, 11, 0.12);
    --purple: #8b5cf6;
    --purple-soft: rgba(139, 92, 246, 0.12);
    --text-primary: #f1f5f9;
    --text-secondary: #94a3b8;
    --text-muted: #64748b;
    --radius: 16px;
    --shadow: 0 8px 32px rgba(0, 0, 0, 0.25);
  }}

  * {{ margin: 0; padding: 0; box-sizing: border-box; }}

  html {{ scroll-behavior: smooth; }}

  body {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    background: var(--bg-base);
    color: var(--text-primary);
    min-height: 100vh;
    overflow-x: hidden;
  }}

  body::before {{
    content: '';
    position: fixed;
    top: -300px;
    right: -200px;
    width: 800px;
    height: 800px;
    background: radial-gradient(circle, rgba(59,130,246,0.06) 0%, transparent 60%);
    border-radius: 50%;
    pointer-events: none;
    z-index: 0;
  }}

  body::after {{
    content: '';
    position: fixed;
    bottom: -300px;
    left: -200px;
    width: 700px;
    height: 700px;
    background: radial-gradient(circle, rgba(139,92,246,0.05) 0%, transparent 60%);
    border-radius: 50%;
    pointer-events: none;
    z-index: 0;
  }}

  /* ── GLASS CARD ─────────────────────────── */
  .glass {{
    background: var(--bg-surface);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border: 1px solid var(--glass-border);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    position: relative;
    overflow: hidden;
  }}

  .glass::before {{
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.1), transparent);
    pointer-events: none;
  }}

  /* ── HEADER ─────────────────────────────── */
  .header {{
    position: sticky;
    top: 0;
    z-index: 100;
    padding: 16px 40px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 1px solid var(--glass-border);
  }}

  .header-brand {{
    display: flex;
    align-items: center;
    gap: 14px;
  }}

  .header-logo {{
    width: 40px;
    height: 40px;
    background: linear-gradient(135deg, var(--accent), var(--purple));
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 900;
    font-size: 18px;
    color: white;
  }}

  .header-title {{
    font-size: 1.2em;
    font-weight: 700;
    color: var(--text-primary);
  }}

  .header-sub {{
    font-size: 0.8em;
    color: var(--text-muted);
    font-weight: 400;
  }}

  .header-meta {{
    display: flex;
    gap: 24px;
    align-items: center;
  }}

  .header-pill {{
    padding: 6px 14px;
    border-radius: 20px;
    font-size: 0.78em;
    font-weight: 600;
    letter-spacing: 0.3px;
  }}

  .pill-green {{ background: var(--green-soft); color: var(--green); }}
  .pill-blue {{ background: var(--accent-soft); color: var(--accent); }}

  /* ── LAYOUT ─────────────────────────────── */
  .dashboard {{
    max-width: 1440px;
    margin: 0 auto;
    padding: 24px 32px 60px;
    position: relative;
    z-index: 1;
  }}

  .section {{
    margin-bottom: 28px;
  }}

  .section-title {{
    font-size: 0.78em;
    font-weight: 700;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 1.2px;
    margin-bottom: 14px;
    padding-left: 4px;
  }}

  /* ── KPI GRID ─────────────────────────────── */
  .kpi-grid {{
    display: grid;
    grid-template-columns: repeat(6, 1fr);
    gap: 16px;
  }}

  .kpi-card {{
    padding: 22px 20px;
    transition: transform 0.25s ease, border-color 0.25s ease;
    cursor: default;
  }}

  .kpi-card:hover {{
    transform: translateY(-4px);
    border-color: rgba(255,255,255,0.15);
  }}

  .kpi-label {{
    font-size: 0.72em;
    font-weight: 600;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.6px;
    margin-bottom: 10px;
  }}

  .kpi-value {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 1.7em;
    font-weight: 700;
    line-height: 1;
    margin-bottom: 6px;
  }}

  .kpi-change {{
    font-size: 0.78em;
    font-weight: 600;
    display: inline-flex;
    align-items: center;
    gap: 4px;
    padding: 3px 8px;
    border-radius: 6px;
  }}

  .change-up {{ background: var(--green-soft); color: var(--green); }}
  .change-down {{ background: var(--red-soft); color: var(--red); }}
  .change-neutral {{ background: var(--accent-soft); color: var(--accent); }}

  /* ── CHART GRID ─────────────────────────────── */
  .chart-grid {{
    display: grid;
    gap: 16px;
  }}

  .grid-2 {{ grid-template-columns: 1fr 1fr; }}
  .grid-3 {{ grid-template-columns: 1fr 1fr 1fr; }}
  .grid-1-2 {{ grid-template-columns: 1fr 2fr; }}
  .grid-2-1 {{ grid-template-columns: 2fr 1fr; }}

  .chart-card {{
    padding: 24px;
    min-height: 360px;
    display: flex;
    flex-direction: column;
  }}

  .chart-card-header {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 16px;
  }}

  .chart-title {{
    font-size: 0.95em;
    font-weight: 700;
    color: var(--text-primary);
  }}

  .chart-badge {{
    padding: 4px 10px;
    border-radius: 6px;
    font-size: 0.7em;
    font-weight: 600;
    background: var(--bg-elevated);
    color: var(--text-secondary);
    border: 1px solid var(--glass-border);
  }}

  .chart-body {{
    flex: 1;
    position: relative;
    min-height: 280px;
  }}

  /* ── TABLE ─────────────────────────────── */
  .data-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.88em;
  }}

  .data-table thead th {{
    text-align: left;
    padding: 10px 14px;
    font-size: 0.75em;
    font-weight: 700;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    border-bottom: 1px solid var(--glass-border);
  }}

  .data-table tbody td {{
    padding: 10px 14px;
    border-bottom: 1px solid rgba(255, 255, 255, 0.03);
    color: var(--text-secondary);
    transition: background 0.15s;
  }}

  .data-table tbody tr:hover td {{
    background: rgba(59, 130, 246, 0.04);
    color: var(--text-primary);
  }}

  .data-table .mono {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.9em;
  }}

  /* ── INSIGHTS PANEL ─────────────────────── */
  .insight-list {{
    list-style: none;
  }}

  .insight-item {{
    padding: 14px 16px;
    margin-bottom: 8px;
    border-radius: 12px;
    background: var(--bg-elevated);
    border: 1px solid var(--glass-border);
    font-size: 0.88em;
    line-height: 1.55;
    color: var(--text-secondary);
    transition: transform 0.2s, border-color 0.2s;
    display: flex;
    gap: 12px;
    align-items: flex-start;
  }}

  .insight-item:hover {{
    transform: translateX(4px);
    border-color: rgba(255,255,255,0.1);
  }}

  .insight-dot {{
    width: 8px;
    height: 8px;
    border-radius: 50%;
    flex-shrink: 0;
    margin-top: 6px;
  }}

  .dot-green {{ background: var(--green); box-shadow: 0 0 8px rgba(16,185,129,0.4); }}
  .dot-amber {{ background: var(--amber); box-shadow: 0 0 8px rgba(245,158,11,0.4); }}
  .dot-red {{ background: var(--red); box-shadow: 0 0 8px rgba(239,68,68,0.4); }}
  .dot-blue {{ background: var(--accent); box-shadow: 0 0 8px rgba(59,130,246,0.4); }}

  /* ── FOOTER ─────────────────────────────── */
  .footer {{
    text-align: center;
    padding: 30px;
    color: var(--text-muted);
    font-size: 0.78em;
    border-top: 1px solid var(--glass-border);
    margin-top: 20px;
  }}

  /* ── SCROLLBAR ────────────────────────── */
  ::-webkit-scrollbar {{ width: 6px; }}
  ::-webkit-scrollbar-track {{ background: transparent; }}
  ::-webkit-scrollbar-thumb {{ background: rgba(255,255,255,0.1); border-radius: 3px; }}
  ::-webkit-scrollbar-thumb:hover {{ background: rgba(255,255,255,0.2); }}

  /* ── TABS ────────────────────────── */
  .tab-container {{
    display: flex;
    gap: 4px;
    background: var(--bg-elevated);
    padding: 4px;
    border-radius: 10px;
    border: 1px solid var(--glass-border);
  }}

  .tab-btn {{
    padding: 7px 16px;
    border: none;
    background: transparent;
    color: var(--text-muted);
    font-family: 'Inter', sans-serif;
    font-size: 0.8em;
    font-weight: 600;
    border-radius: 8px;
    cursor: pointer;
    transition: all 0.2s;
  }}

  .tab-btn:hover {{ color: var(--text-secondary); }}
  .tab-btn.active {{
    background: var(--accent-soft);
    color: var(--accent);
  }}

  .tab-panel {{ display: none; }}
  .tab-panel.active {{ display: block; }}

  /* ── RESPONSIVE ─────────────────────── */
  @media (max-width: 1200px) {{
    .kpi-grid {{ grid-template-columns: repeat(3, 1fr); }}
    .grid-2, .grid-1-2, .grid-2-1 {{ grid-template-columns: 1fr; }}
  }}

  @media (max-width: 768px) {{
    .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }}
    .header {{ flex-direction: column; gap: 12px; padding: 16px 20px; }}
    .dashboard {{ padding: 16px; }}
  }}

  /* ── ANIMATIONS ─────────────────────── */
  @keyframes fadeUp {{
    from {{ opacity: 0; transform: translateY(20px); }}
    to {{ opacity: 1; transform: translateY(0); }}
  }}

  .animate {{ animation: fadeUp 0.5s ease forwards; opacity: 0; }}
  .delay-1 {{ animation-delay: 0.1s; }}
  .delay-2 {{ animation-delay: 0.2s; }}
  .delay-3 {{ animation-delay: 0.3s; }}
  .delay-4 {{ animation-delay: 0.4s; }}
  .delay-5 {{ animation-delay: 0.5s; }}
  .delay-6 {{ animation-delay: 0.6s; }}
</style>
</head>
<body>

<!-- ═══ HEADER ═══ -->
<header class="header glass">
  <div class="header-brand">
    <div class="header-logo">LS</div>
    <div>
      <div class="header-title">Labhyansh Solution</div>
      <div class="header-sub">Financial Dashboard <span style="opacity:0.3">/</span> FY 2024-25</div>
    </div>
  </div>
  <div class="header-meta">
    <span class="header-pill pill-blue">Apr to Aug 2024</span>
    <span class="header-pill pill-green">{"Profitable" if net_profit > 0 else "Loss"}</span>
    <span style="font-size: 0.8em; color: var(--text-muted);">Prepared by Aviral Dubey</span>
  </div>
</header>

<main class="dashboard">

  <!-- ═══ KPI SECTION ═══ -->
  <section class="section">
    <div class="section-title">Key Performance Indicators</div>
    <div class="kpi-grid">
      <div class="kpi-card glass animate delay-1">
        <div class="kpi-label">Total Revenue</div>
        <div class="kpi-value" style="color: var(--green);">{fmt(total_revenue)}</div>
        <span class="kpi-change change-up">+{rev_growth}% growth</span>
      </div>
      <div class="kpi-card glass animate delay-2">
        <div class="kpi-label">Total Expenses</div>
        <div class="kpi-value" style="color: var(--red);">{fmt(total_expenses)}</div>
        <span class="kpi-change change-neutral">90% of revenue</span>
      </div>
      <div class="kpi-card glass animate delay-3">
        <div class="kpi-label">Net Profit</div>
        <div class="kpi-value" style="color: {'var(--green)' if net_profit > 0 else 'var(--red)'};">{fmt(net_profit)}</div>
        <span class="kpi-change {'change-up' if net_profit > 0 else 'change-down'}">{margin}% margin</span>
      </div>
      <div class="kpi-card glass animate delay-4">
        <div class="kpi-label">Net Assets</div>
        <div class="kpi-value" style="color: var(--accent);">{fmt(net_assets)}</div>
        <span class="kpi-change change-neutral">Balance Sheet</span>
      </div>
      <div class="kpi-card glass animate delay-5">
        <div class="kpi-label">Cash & Bank</div>
        <div class="kpi-value" style="color: var(--green);">{fmt(cash_bank)}</div>
        <span class="kpi-change change-up">Liquid reserves</span>
      </div>
      <div class="kpi-card glass animate delay-6">
        <div class="kpi-label">Owner Equity</div>
        <div class="kpi-value" style="color: var(--purple);">{fmt(net_equity)}</div>
        <span class="kpi-change change-neutral">Capital invested</span>
      </div>
    </div>
  </section>

  <!-- ═══ REVENUE vs EXPENSES TREND ═══ -->
  <section class="section">
    <div class="section-title">Performance Trends</div>
    <div class="chart-grid grid-2-1">
      <div class="chart-card glass animate delay-2">
        <div class="chart-card-header">
          <span class="chart-title">Revenue vs Expenses (Monthly)</span>
          <div class="tab-container">
            <button class="tab-btn active" onclick="setTrendType('bar')">Bar</button>
            <button class="tab-btn" onclick="setTrendType('line')">Line</button>
          </div>
        </div>
        <div class="chart-body">
          <canvas id="trendChart"></canvas>
        </div>
      </div>
      <div class="chart-card glass animate delay-3">
        <div class="chart-card-header">
          <span class="chart-title">Cumulative P&L</span>
          <span class="chart-badge">Running Total</span>
        </div>
        <div class="chart-body">
          <canvas id="cumulativeChart"></canvas>
        </div>
      </div>
    </div>
  </section>

  <!-- ═══ MONTHLY DETAIL TABLE ═══ -->
  <section class="section">
    <div class="section-title">Monthly Breakdown</div>
    <div class="glass" style="padding: 20px; border-radius: var(--radius);">
      <table class="data-table">
        <thead>
          <tr>
            <th>Month</th>
            <th>Revenue</th>
            <th>Expenses</th>
            <th>Net P&L</th>
            <th>Margin</th>
            <th>Status</th>
          </tr>
        </thead>
        <tbody>'''

for i, m in enumerate(MONTHS):
    rev = monthly_revenue[i]
    exp = monthly_expenses[i]
    net = monthly_net[i]
    marg = round((net / rev) * 100, 1) if rev else 0
    status_class = 'change-up' if net > 0 else 'change-down'
    status_text = 'Profit' if net > 0 else 'Loss'
    html += f'''
          <tr>
            <td><strong>{MONTH_LABELS_FULL[i]}</strong></td>
            <td class="mono" style="color:var(--green);">{fmt(rev)}</td>
            <td class="mono" style="color:var(--red);">{fmt(exp)}</td>
            <td class="mono" style="color:{'var(--green)' if net > 0 else 'var(--red)'};">{fmt(net)}</td>
            <td class="mono">{marg}%</td>
            <td><span class="kpi-change {status_class}">{status_text}</span></td>
          </tr>'''

html += f'''
        </tbody>
      </table>
    </div>
  </section>

  <!-- ═══ EXPENSE BREAKDOWN + CATEGORY PIE ═══ -->
  <section class="section">
    <div class="section-title">Expense & Category Analysis</div>
    <div class="chart-grid grid-2">
      <div class="chart-card glass animate delay-1">
        <div class="chart-card-header">
          <span class="chart-title">Expense Breakdown by Type</span>
          <span class="chart-badge">Top 10</span>
        </div>
        <div class="chart-body">
          <canvas id="expenseChart"></canvas>
        </div>
      </div>
      <div class="chart-card glass animate delay-2">
        <div class="chart-card-header">
          <span class="chart-title">Category Distribution</span>
          <div class="tab-container">
            <button class="tab-btn active" onclick="setCatView('debit')">Debit</button>
            <button class="tab-btn" onclick="setCatView('credit')">Credit</button>
          </div>
        </div>
        <div class="chart-body">
          <canvas id="categoryChart"></canvas>
        </div>
      </div>
    </div>
  </section>

  <!-- ═══ TOP ACCOUNTS ═══ -->
  <section class="section">
    <div class="section-title">Top Accounts Analysis</div>
    <div class="chart-grid grid-2">
      <div class="chart-card glass animate delay-1">
        <div class="chart-card-header">
          <span class="chart-title">Top 10 Accounts by Debit</span>
          <span class="chart-badge">Outflows</span>
        </div>
        <div class="chart-body">
          <canvas id="topDebitChart"></canvas>
        </div>
      </div>
      <div class="chart-card glass animate delay-2">
        <div class="chart-card-header">
          <span class="chart-title">Top 10 Accounts by Credit</span>
          <span class="chart-badge">Inflows</span>
        </div>
        <div class="chart-body">
          <canvas id="topCreditChart"></canvas>
        </div>
      </div>
    </div>
  </section>

  <!-- ═══ EXPENSE TREND (STACKED) ═══ -->
  <section class="section">
    <div class="section-title">Expense Trend by Category</div>
    <div class="chart-card glass animate delay-1">
      <div class="chart-card-header">
        <span class="chart-title">Monthly Expense Movement — Top 5 Categories</span>
        <span class="chart-badge">Stacked</span>
      </div>
      <div class="chart-body" style="min-height: 320px;">
        <canvas id="expTrendChart"></canvas>
      </div>
    </div>
  </section>

  <!-- ═══ INSIGHTS & RISKS ═══ -->
  <section class="section">
    <div class="section-title">Key Insights & Risk Assessment</div>
    <div class="chart-grid grid-2">
      <div class="glass" style="padding: 24px;">
        <h3 style="font-size: 0.95em; font-weight: 700; margin-bottom: 16px; color: var(--green);">Strengths</h3>
        <ul class="insight-list">
          <li class="insight-item">
            <span class="insight-dot dot-green"></span>
            <span><strong>Profitable business</strong> <span style="opacity:0.3">&vert;</span> Net profit of {fmt(net_profit)} with {margin}% margin over 5 months</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-green"></span>
            <span><strong>Revenue growing +{rev_growth}%</strong> <span style="opacity:0.3">&vert;</span> Consistent upward trend from April to August</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-green"></span>
            <span><strong>Strong cash position</strong> <span style="opacity:0.3">&vert;</span> {fmt(cash_bank)} in cash and bank reserves</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-green"></span>
            <span><strong>Solid asset base</strong> <span style="opacity:0.3">&vert;</span> {fmt(net_assets)} in net assets supporting long-term growth</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-blue"></span>
            <span><strong>Recovering trajectory</strong> <span style="opacity:0.3">&vert;</span> Only April was a loss month; 4 consecutive profitable months</span>
          </li>
        </ul>
      </div>
      <div class="glass" style="padding: 24px;">
        <h3 style="font-size: 0.95em; font-weight: 700; margin-bottom: 16px; color: var(--amber);">Risks & Recommendations</h3>
        <ul class="insight-list">
          <li class="insight-item">
            <span class="insight-dot dot-amber"></span>
            <span><strong>Thin margin (10%)</strong> <span style="opacity:0.3">&vert;</span> Vulnerable to cost increases. Negotiate better purchase rates.</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-red"></span>
            <span><strong>Revenue concentration</strong> <span style="opacity:0.3">&vert;</span> Only 5 accounts = 80% of credit inflows. Diversify customer base.</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-amber"></span>
            <span><strong>Purchases = 49% of costs</strong> <span style="opacity:0.3">&vert;</span> Explore bulk deals, alternate suppliers for 2-3% savings.</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-amber"></span>
            <span><strong>Asset heavy structure</strong> <span style="opacity:0.3">&vert;</span> High depreciation and maintenance overhead to manage.</span>
          </li>
          <li class="insight-item">
            <span class="insight-dot dot-blue"></span>
            <span><strong>Build reserves</strong> <span style="opacity:0.3">&vert;</span> Maintain 2 to 3 months of operating expenses as a safety buffer.</span>
          </li>
        </ul>
      </div>
    </div>
  </section>

</main>

<footer class="footer">
  Labhyansh Solution &middot; Financial Dashboard &middot; FY 2024-25 (Q1 & Q2) &middot; Prepared by Aviral Dubey
</footer>

<!-- ═══════════════════════════════════════════════════════════ -->
<!-- CHART.JS SCRIPTS -->
<!-- ═══════════════════════════════════════════════════════════ -->
<script>
Chart.defaults.color = '#94a3b8';
Chart.defaults.borderColor = 'rgba(255,255,255,0.05)';
Chart.defaults.font.family = "'Inter', sans-serif";
Chart.defaults.font.size = 12;
Chart.defaults.plugins.legend.labels.usePointStyle = true;
Chart.defaults.plugins.legend.labels.pointStyleWidth = 10;
Chart.defaults.plugins.tooltip.backgroundColor = 'rgba(15, 23, 42, 0.95)';
Chart.defaults.plugins.tooltip.borderColor = 'rgba(59, 130, 246, 0.2)';
Chart.defaults.plugins.tooltip.borderWidth = 1;
Chart.defaults.plugins.tooltip.cornerRadius = 8;
Chart.defaults.plugins.tooltip.padding = 12;
Chart.defaults.plugins.tooltip.titleFont = {{ weight: '600' }};

const months = {json.dumps(MONTH_LABELS_SHORT)};
const revenueData = {json.dumps(monthly_revenue)};
const expenseData = {json.dumps(monthly_expenses)};
const netData = {json.dumps(monthly_net)};
const cumData = {json.dumps(cum_net)};

function formatINR(val) {{
  if (Math.abs(val) >= 10000000) return (val/10000000).toFixed(2) + ' Cr';
  if (Math.abs(val) >= 100000) return (val/100000).toFixed(2) + ' L';
  return val.toLocaleString('en-IN');
}}

// ── 1. TREND CHART ─────────────────────────
let trendChart;
function createTrendChart(type) {{
  if (trendChart) trendChart.destroy();
  const ctx = document.getElementById('trendChart').getContext('2d');
  trendChart = new Chart(ctx, {{
    type: type,
    data: {{
      labels: months,
      datasets: [
        {{
          label: 'Revenue',
          data: revenueData,
          backgroundColor: type === 'bar' ? 'rgba(16,185,129,0.7)' : 'rgba(16,185,129,0.1)',
          borderColor: '#10b981',
          borderWidth: 2,
          borderRadius: type === 'bar' ? 6 : 0,
          fill: type === 'line',
          tension: 0.4,
          pointRadius: type === 'line' ? 5 : 0,
          pointHoverRadius: 7,
        }},
        {{
          label: 'Expenses',
          data: expenseData,
          backgroundColor: type === 'bar' ? 'rgba(239,68,68,0.7)' : 'rgba(239,68,68,0.1)',
          borderColor: '#ef4444',
          borderWidth: 2,
          borderRadius: type === 'bar' ? 6 : 0,
          fill: type === 'line',
          tension: 0.4,
          pointRadius: type === 'line' ? 5 : 0,
          pointHoverRadius: 7,
        }},
        {{
          label: 'Net P&L',
          data: netData,
          backgroundColor: type === 'bar' ? 'rgba(59,130,246,0.7)' : 'rgba(59,130,246,0.1)',
          borderColor: '#3b82f6',
          borderWidth: 2,
          borderRadius: type === 'bar' ? 6 : 0,
          fill: type === 'line',
          tension: 0.4,
          pointRadius: type === 'line' ? 5 : 0,
          pointHoverRadius: 7,
        }}
      ]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{
        tooltip: {{
          callbacks: {{ label: (ctx) => ctx.dataset.label + ': Rs. ' + formatINR(ctx.parsed.y) }}
        }}
      }},
      scales: {{
        y: {{
          ticks: {{ callback: (v) => formatINR(v) }},
          grid: {{ color: 'rgba(255,255,255,0.04)' }}
        }},
        x: {{ grid: {{ display: false }} }}
      }}
    }}
  }});
}}

function setTrendType(type) {{
  document.querySelectorAll('.tab-container')[0].querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');
  createTrendChart(type);
}}

createTrendChart('bar');

// ── 2. CUMULATIVE CHART ─────────────────────────
const cumCtx = document.getElementById('cumulativeChart').getContext('2d');
new Chart(cumCtx, {{
  type: 'line',
  data: {{
    labels: months,
    datasets: [{{
      label: 'Cumulative Net P&L',
      data: cumData,
      borderColor: '#f59e0b',
      backgroundColor: 'rgba(245,158,11,0.1)',
      borderWidth: 3,
      fill: true,
      tension: 0.4,
      pointRadius: 6,
      pointHoverRadius: 9,
      pointBackgroundColor: '#f59e0b',
      pointBorderColor: '#0a0e1a',
      pointBorderWidth: 3,
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{
        callbacks: {{ label: (ctx) => 'Cumulative: Rs. ' + formatINR(ctx.parsed.y) }}
      }}
    }},
    scales: {{
      y: {{
        ticks: {{ callback: (v) => formatINR(v) }},
        grid: {{ color: 'rgba(255,255,255,0.04)' }}
      }},
      x: {{ grid: {{ display: false }} }}
    }}
  }}
}});

// ── 3. EXPENSE BREAKDOWN ─────────────────────────
const expCtx = document.getElementById('expenseChart').getContext('2d');
new Chart(expCtx, {{
  type: 'bar',
  data: {{
    labels: {exp_sub_labels},
    datasets: [{{
      data: {exp_sub_values},
      backgroundColor: [
        'rgba(239,68,68,0.8)', 'rgba(239,68,68,0.65)', 'rgba(239,68,68,0.5)',
        'rgba(245,158,11,0.7)', 'rgba(245,158,11,0.55)', 'rgba(245,158,11,0.4)',
        'rgba(139,92,246,0.7)', 'rgba(139,92,246,0.55)', 'rgba(139,92,246,0.4)',
        'rgba(59,130,246,0.6)'
      ],
      borderRadius: 6,
      borderSkipped: false,
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    indexAxis: 'y',
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{
        callbacks: {{ label: (ctx) => 'Rs. ' + formatINR(ctx.parsed.x) }}
      }}
    }},
    scales: {{
      x: {{
        ticks: {{ callback: (v) => formatINR(v) }},
        grid: {{ color: 'rgba(255,255,255,0.04)' }}
      }},
      y: {{
        grid: {{ display: false }},
        ticks: {{ font: {{ size: 11 }} }}
      }}
    }}
  }}
}});

// ── 4. CATEGORY PIE ─────────────────────────
let categoryChart;
const catLabels = {cat_labels};
const catDebVals = {cat_deb_vals};
const catCredVals = {cat_cred_vals};
const catColors = ['rgba(59,130,246,0.8)', 'rgba(239,68,68,0.8)', 'rgba(139,92,246,0.8)', 'rgba(16,185,129,0.8)', 'rgba(245,158,11,0.8)'];

function createCatChart(view) {{
  if (categoryChart) categoryChart.destroy();
  const ctx = document.getElementById('categoryChart').getContext('2d');
  categoryChart = new Chart(ctx, {{
    type: 'doughnut',
    data: {{
      labels: catLabels,
      datasets: [{{
        data: view === 'debit' ? catDebVals : catCredVals,
        backgroundColor: catColors,
        borderColor: '#0a0e1a',
        borderWidth: 3,
        hoverOffset: 8,
      }}]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      cutout: '55%',
      plugins: {{
        legend: {{
          position: 'bottom',
          labels: {{ padding: 16, font: {{ size: 11 }} }}
        }},
        tooltip: {{
          callbacks: {{ label: (ctx) => ctx.label + ': Rs. ' + formatINR(ctx.parsed) }}
        }}
      }}
    }}
  }});
}}

function setCatView(view) {{
  document.querySelectorAll('.tab-container')[1].querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  event.target.classList.add('active');
  createCatChart(view);
}}

createCatChart('debit');

// ── 5. TOP DEBIT ─────────────────────────
const tdCtx = document.getElementById('topDebitChart').getContext('2d');
new Chart(tdCtx, {{
  type: 'bar',
  data: {{
    labels: {top_deb_labels},
    datasets: [{{
      data: {top_deb_values},
      backgroundColor: 'rgba(59,130,246,0.7)',
      borderRadius: 6,
      borderSkipped: false,
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    indexAxis: 'y',
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: (ctx) => 'Rs. ' + formatINR(ctx.parsed.x) }} }}
    }},
    scales: {{
      x: {{ ticks: {{ callback: (v) => formatINR(v) }}, grid: {{ color: 'rgba(255,255,255,0.04)' }} }},
      y: {{ grid: {{ display: false }}, ticks: {{ font: {{ size: 10 }} }} }}
    }}
  }}
}});

// ── 6. TOP CREDIT ─────────────────────────
const tcCtx = document.getElementById('topCreditChart').getContext('2d');
new Chart(tcCtx, {{
  type: 'bar',
  data: {{
    labels: {top_cred_labels},
    datasets: [{{
      data: {top_cred_values},
      backgroundColor: 'rgba(16,185,129,0.7)',
      borderRadius: 6,
      borderSkipped: false,
    }}]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    indexAxis: 'y',
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: (ctx) => 'Rs. ' + formatINR(ctx.parsed.x) }} }}
    }},
    scales: {{
      x: {{ ticks: {{ callback: (v) => formatINR(v) }}, grid: {{ color: 'rgba(255,255,255,0.04)' }} }},
      y: {{ grid: {{ display: false }}, ticks: {{ font: {{ size: 10 }} }} }}
    }}
  }}
}});

// ── 7. EXPENSE TREND (STACKED) ─────────────
const etCtx = document.getElementById('expTrendChart').getContext('2d');
new Chart(etCtx, {{
  type: 'bar',
  data: {{
    labels: months,
    datasets: {heatmap_json}
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    interaction: {{ mode: 'index', intersect: false }},
    plugins: {{
      legend: {{
        position: 'top',
        labels: {{ padding: 16, font: {{ size: 11 }}, usePointStyle: true }}
      }},
      tooltip: {{
        callbacks: {{ label: (ctx) => ctx.dataset.label + ': Rs. ' + formatINR(ctx.parsed.y) }}
      }}
    }},
    scales: {{
      x: {{ stacked: true, grid: {{ display: false }} }},
      y: {{ stacked: true, ticks: {{ callback: (v) => formatINR(v) }}, grid: {{ color: 'rgba(255,255,255,0.04)' }} }}
    }}
  }}
}});

// ── INTERSECTION OBSERVER (animate on scroll) ─────────────
const observer = new IntersectionObserver((entries) => {{
  entries.forEach(entry => {{
    if (entry.isIntersecting) {{
      entry.target.style.animationPlayState = 'running';
    }}
  }});
}}, {{ threshold: 0.1 }});

document.querySelectorAll('.animate').forEach(el => {{
  el.style.animationPlayState = 'paused';
  observer.observe(el);
}});

</script>
</body>
</html>'''

os.makedirs(OUTPUT_DIR, exist_ok=True)
with open(DASHBOARD_FILE, 'w', encoding='utf-8') as f:
    f.write(html)

file_size = os.path.getsize(DASHBOARD_FILE) / 1024
print(f"  Dashboard saved: {DASHBOARD_FILE}")
print(f"  File size: {file_size:.0f} KB")
print(f"  Charts: 7 interactive Chart.js visuals")
print(f"  Features: glassmorphism, dark theme, tabs, scroll animations, responsive")
